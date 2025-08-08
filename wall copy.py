import pandas as pd
import math
import itertools
import re
import plotly.graph_objects as go

# File paths
bom_file = "/Users/craignyabvure/Desktop/Python Test Projects AGI/Trailer Loading/BOM Wall Parts.xlsx"
solver_file = "/Users/craignyabvure/Desktop/Python Test Projects AGI/Trailer Loading/Solver (2).xlsx"
output_file = "/Users/craignyabvure/Desktop/Python Test Projects AGI/Trailer Loading/Loading_Plan_Output.xlsx"

# Trailer specs (in inches for dimensions)
trailers = {
    'Tandem': {'length': 624, 'width': 96, 'height': 20, 'max_volume': 3480.32, 'max_weight': 470000},
    'Flat Deck 53': {'length': 636, 'width': 102, 'height': 20, 'max_volume': 3842.85, 'max_weight': 46000},
    'Super B': {'length': 720, 'width': 102, 'height': 20, 'max_volume': 4350.39, 'max_weight': 92000}
}

# Cost map for trailers
cost_map = {'Tandem': 5000, 'Flat Deck 53': 6500, 'Super B': 7500}

# Country restriction configuration
country = "Canada"  # Change this value as needed
restricted = {"america", "us", "usa", "united states", "united states of america"}

#
# Sheet dimensions and volume
sheet_length = 116.5
sheet_width = 45.5
sheet_height = 0.25
sheet_volume = 0.767  # cubic feet

# --- Constants for wall stacking ---
gap_between_stacks = 3.5  # Average 3-4 inch gap (inches)
gap_between_stacks_length = 2.5  # inches
stack_height_limit_pct = 0.92

# Load Excel data
bom_df = pd.read_excel(bom_file)
solver_df = pd.read_excel(solver_file)

# Normalize columns
bom_df.columns = bom_df.columns.str.strip().str.lower()
solver_df.columns = solver_df.columns.str.strip().str.lower()

# Rename relevant columns for consistency
bom_df.rename(columns={'assembly': 'assembly', 'material': 'material', 'quantity': 'quantity'}, inplace=True)
solver_df.rename(columns={'material': 'material', 'material desc': 'description', 'matl grou': 'group', 'net weight': 'weight'}, inplace=True)

# Ensure material columns are strings and lowercase before merge
bom_df['material'] = bom_df['material'].astype(str).str.strip().str.lower()
solver_df['material'] = solver_df['material'].astype(str).str.strip().str.lower()

# Merge BOM with Solver on 'material'
merged_df = pd.merge(bom_df, solver_df, on='material', how='left')

# Extract numeric part before 'B' only for valid roof materials
def extract_bin(material_code):
    match = re.match(r"(\d+)b", material_code.lower())
    return match.group(1) if match else None

solver_df['roof_bin'] = solver_df.apply(
    lambda row: extract_bin(row['material']) if row['group'] == 'roof' else None,
    axis=1
)

# Identify unmatched BOM materials not found in the solver
unmatched_bom = bom_df[~bom_df['material'].isin(solver_df['material'])].copy()
unmatched_bom['group'] = 'unknown'
unmatched_bom['weight'] = None
unmatched_bom['reason'] = 'Material not found in solver database'
unmatched_bom = unmatched_bom[['assembly', 'material', 'group', 'quantity', 'weight', 'reason']]

# Normalize 'assembly' after merge
merged_df['assembly'] = merged_df['assembly'].astype(str).str.strip().str.lower()

# Normalize 'group' column
merged_df['group'] = merged_df['group'].astype(str).str.strip().str.lower()
# Normalize 'weight' column to numeric
merged_df['weight'] = pd.to_numeric(merged_df['weight'], errors='coerce')

# Validate essential columns exist
required_columns = ['assembly', 'material', 'quantity', 'group', 'weight']
for col in required_columns:
    if col not in merged_df.columns:
        raise ValueError(f"Required column '{col}' not found in the merged data.")

# Group the merged data to eliminate repeats
merged_df = merged_df.groupby(['assembly', 'material', 'group', 'weight'], as_index=False)['quantity'].sum()

 # Now filter for loadable items only (e.g., wall, roof) and valid weight

# --- ROOF SHEET DIMENSIONS HANDLING ---
def prepare_roof_items(merged_df, roof_dims_path):
    # Load dimensions sheet
    roof_df = pd.read_excel(roof_dims_path)
    roof_df.columns = roof_df.columns.str.strip().str.lower()
    
    # Normalize material columns
    roof_df['material'] = roof_df['material'].astype(str).str.strip().str.lower()
    
    # Extract relevant roof materials from merged_df
    roof_items = merged_df[merged_df['group'] == 'roof'].copy()
    # Extract bin number
    roof_items['roof_bin'] = roof_items['material'].apply(lambda mat: extract_bin(mat))
    roof_items = roof_items[roof_items['roof_bin'].notna()]

    # Merge roof dimensions on bin number
    roof_items = roof_items.merge(roof_df, left_on='roof_bin', right_on='material', how='left', suffixes=('', '_roofdim'))

    # Ensure dimensional columns exist
    for col in ['length', 'width', 'height']:
        if col not in roof_items.columns:
            raise ValueError(f"Missing '{col}' in Roof Sheet Dimensions file.")

    # Calculate volume (convert inches³ to ft³)
    roof_items['length'] = pd.to_numeric(roof_items['length'], errors='coerce')
    roof_items['width'] = pd.to_numeric(roof_items['width'], errors='coerce')
    roof_items['height'] = pd.to_numeric(roof_items['height'], errors='coerce')
    roof_items['volume'] = (roof_items['length'] * roof_items['width'] * roof_items['height']) / 1728

    # Drop items without valid dimensions or weights
    roof_items.dropna(subset=['length', 'width', 'height', 'volume', 'weight'], inplace=True)

    return roof_items

# Prepare wall and roof items
roof_dims_path = "/Users/craignyabvure/Desktop/Python Test Projects AGI/Trailer Loading/Roof Sheet Dimensions.xlsx"
roof_items_df = prepare_roof_items(merged_df, roof_dims_path)

loadable_groups = ['wall']  # We'll manually merge roof after processing
wall_items_df = merged_df[merged_df['group'].isin(loadable_groups)].copy()
wall_items_df['weight'] = pd.to_numeric(wall_items_df['weight'], errors='coerce')
wall_items_df['quantity'] = pd.to_numeric(wall_items_df['quantity'], errors='coerce').fillna(0).astype(int)

# Combine both wall and roof
all_loadable_items_df = pd.concat([wall_items_df, roof_items_df], ignore_index=True)
all_loadable_items_df = all_loadable_items_df.sort_values(by=['weight', 'quantity'], ascending=[False, False]).reset_index(drop=True)

# Prepare non-wall or missing weight items for unloaded sheet later
non_wall_or_missing_weight_df = merged_df[
    (merged_df['group'] != 'wall') | (pd.isna(merged_df['weight']))
].copy()

# Add unmatched BOM rows to non_wall_or_missing_weight_df
non_wall_or_missing_weight_df = pd.concat([
    non_wall_or_missing_weight_df.assign(reason='Material not loaded (non-wall or missing weight)'),
    unmatched_bom
], ignore_index=True)

# Function to simulate loading materials into a single trailer
def try_fit_trailer(trailer_name, trailer_specs, materials_df):
    trailer_log = []
    loaded_set = set()
    total_volume = 0
    total_weight = 0
    total_height = 0
    stack_count = 0

    first_sheet = True
    sheets_per_row = trailer_specs['width'] // sheet_width
    max_rows = trailer_specs['length'] // sheet_length
    max_stack_height = trailer_specs['height']

    for idx, row in materials_df.iterrows():
        quantity = int(row['quantity'])
        weight_per_sheet = row['weight']
        loaded_count = 0
        for i in range(quantity):
            if first_sheet:
                added_height = 9.59
                added_volume = 29.42
                first_sheet = False
            else:
                added_height = sheet_height
                added_volume = sheet_volume
            if total_height + added_height > max_stack_height or \
               total_volume + added_volume > trailer_specs['max_volume'] or \
               total_weight + weight_per_sheet > trailer_specs['max_weight']:
                break
            stack_count += 1
            total_height += added_height
            total_volume += added_volume
            total_weight += weight_per_sheet
            loaded_count += 1
            row_number = stack_count // sheets_per_row
            if row_number > max_rows:
                break
        if loaded_count > 0:
            trailer_log.append((row['assembly'], row['material'], row['group'], "Loaded", loaded_count, weight_per_sheet * loaded_count, loaded_count * sheet_volume))
            loaded_set.add((row['assembly'], row['material']))
        else:
            trailer_log.append((row['assembly'], row['material'], row['group'], "Not loaded", quantity, weight_per_sheet * quantity, quantity * sheet_volume))

    return trailer_log, total_volume, total_weight, total_height, loaded_set

# Function to simulate loading for a combination of trailers (1 to 3 distinct trailers)
# Refactored: Fill each trailer one at a time, loading as many full material quantities as possible before moving to the next trailer.
def try_fit_trailers_combination(trailer_names, trailers_specs, materials_df):
    # Sort trailers by cost (lowest first)
    sorted_trailers = sorted(trailer_names, key=lambda t: cost_map[t])
    # Initialize trailer states
    trailers_state = {}
    for tname in sorted_trailers:
        trailers_state[tname] = {
            'specs': trailers_specs[tname],
            'log': [],
            'total_volume': 0,
            'total_weight': 0,
            'total_height': 0,
            'loaded_set': set(),
            'stack_count': 0,
            'first_sheet': True,
            'remaining_width': trailers_specs[tname]['width'],  # Start with full width
            'remaining_length': trailers_specs[tname]['length'],  # Start with full length
            'stack_heights': [],
        }

    # Track remaining quantity for each material by index
    remaining_quantities = {idx: int(row['quantity']) for idx, row in materials_df.iterrows()}

    # For each trailer: fill to capacity by stacking columns side-by-side (width), then new row (length)
    for tname in sorted_trailers:
        state = trailers_state[tname]
        specs = state['specs']
        sheets_per_row = specs['width'] // sheet_width
        max_rows = specs['length'] // sheet_length
        max_stack_height = specs['height']
        # Stacking logic: fill width first, then length
        while state['remaining_length'] >= sheet_length:
            state['remaining_width'] = specs['width']
            while state['remaining_width'] >= sheet_width:
                stack_max_height = max_stack_height * stack_height_limit_pct
                temp_total_height = 0
                temp_total_volume = state['total_volume']
                temp_total_weight = state['total_weight']
                temp_stack_count = 0
                temp_first_sheet = True
                loaded_any = False
                # Allow partial loads per assembly if not all components fit
                for assembly, group_df in materials_df.groupby('assembly'):
                    group_df_sorted = group_df.sort_values(by=['quantity', 'weight'], ascending=[False, False])
                    for idx, row in group_df_sorted.iterrows():
                        quantity_left = remaining_quantities.get(idx, 0)
                        if quantity_left <= 0:
                            continue
                        weight_per_sheet = row['weight']
                        loaded_count = 0
                        for i in range(quantity_left):
                            added_height = 9.59 if temp_first_sheet else sheet_height
                            if row['group'] == 'wall':
                                added_volume = 29.42 if temp_first_sheet else sheet_volume
                            else:
                                added_volume = row['volume']
                            temp_first_sheet = False
                            if (
                                temp_total_height + added_height > stack_max_height or
                                temp_total_volume + added_volume > specs['max_volume'] or
                                temp_total_weight + weight_per_sheet > specs['max_weight']
                            ):
                                break
                            temp_stack_count += 1
                            temp_total_height += added_height
                            temp_total_volume += added_volume
                            temp_total_weight += weight_per_sheet
                            loaded_count += 1
                        if loaded_count > 0:
                            # Update trailer state for this stack
                            state['total_volume'] = temp_total_volume
                            state['total_weight'] = temp_total_weight
                            state['log'].append((
                                row['assembly'], row['material'], row['group'], "Loaded",
                                loaded_count, weight_per_sheet * loaded_count, loaded_count * (sheet_volume if row['group'] == 'wall' else row['volume']),
                                sheet_length if row['group'] == 'wall' else row['length'],
                                sheet_width if row['group'] == 'wall' else row['width']
                            ))
                            state['loaded_set'].add((row['assembly'], row['material']))
                            remaining_quantities[idx] -= loaded_count
                            loaded_any = True
                # Ensure stack height is recorded if stack was built
                if temp_stack_count > 0:
                    state['stack_heights'].append(temp_total_height)
                if not loaded_any:
                    break
                state['remaining_width'] -= sheet_width + gap_between_stacks
            state['remaining_length'] -= sheet_length + gap_between_stacks_length

    # After all trailers, collect unloaded items
    unloaded_items = []
    for idx, row in materials_df.iterrows():
        quantity_left = remaining_quantities.get(idx, 0)
        if quantity_left > 0:
            weight_per_sheet = row['weight']
            unloaded_items.append((
                row['assembly'], row['material'], row['group'], "Not loaded",
                quantity_left, weight_per_sheet * quantity_left, quantity_left * (sheet_volume if row['group'] == 'wall' else row['volume']),
                sheet_length if row['group'] == 'wall' else row['length'],
                sheet_width if row['group'] == 'wall' else row['width']
            ))
    # Add non-wall or missing weight items to unloaded list
    for idx, row in non_wall_or_missing_weight_df.iterrows():
        unloaded_items.append((
            row['assembly'], row['material'], row['group'],
            row['reason'],
            row.get('quantity', None), row.get('weight', None), None,
            None, None
        ))
    # Check if all items loaded
    all_loaded = (len(unloaded_items) == 0 or all(
        (r[3].startswith("Material not loaded") for r in unloaded_items)
    ))
    return trailers_state, unloaded_items, all_loaded

#
# Generate all combinations of 1 to 3 distinct trailers, no repeats, with country restriction support
if country.lower() in restricted:
    all_trailer_names = [t for t in trailers if t != "Super B"]
else:
    all_trailer_names = list(trailers.keys())
all_combinations = []
for r in range(1, 4):
    combis = list(itertools.combinations(all_trailer_names, r))
    all_combinations.extend(combis)

best_config = None
best_config_cost = None
best_config_trailer_count = None
best_config_state = None
best_config_unloaded = None
best_config_loaded_count = -1

# Evaluate each combination
for combi in all_combinations:
    state, unloaded, all_loaded = try_fit_trailers_combination(combi, trailers, all_loadable_items_df)
    total_cost = sum(cost_map[tname] for tname in combi)
    trailer_count = len(combi)
    # Count loaded items for partial solutions
    loaded_count = sum(
        1 for tname in combi for entry in state[tname]['log'] if entry[3] == "Loaded"
    )
    # Select the best config: all loaded and lowest cost, or most loaded if none fit all
    if all_loaded:
        if best_config is None or total_cost < best_config_cost:
            best_config = combi
            best_config_cost = total_cost
            best_config_trailer_count = trailer_count
            best_config_state = state
            best_config_unloaded = unloaded
            best_config_loaded_count = loaded_count
    else:
        if best_config is None or (not best_config_state) or (
            loaded_count > best_config_loaded_count or
            (loaded_count == best_config_loaded_count and (total_cost < (best_config_cost or float('inf'))))
        ):
            best_config = combi
            best_config_cost = total_cost
            best_config_trailer_count = trailer_count
            best_config_state = state
            best_config_unloaded = unloaded
            best_config_loaded_count = loaded_count

from collections import defaultdict

# Prepare Loading Plan log with trailer assignment
log_with_trailer = []
# maps (assembly, material, group, trailer) → [qty, weight, volume]
log_aggregator = defaultdict(lambda: [0, 0.0, 0.0])

# After loading, aggregate logs by trailer assignment for correct quantity calculation
for tname in best_config:
    for entry in best_config_state[tname]['log']:
        if entry[3] == "Loaded":
            key = (entry[0], entry[1], entry[2], tname)
            log_aggregator[key][0] += entry[4]
            log_aggregator[key][1] += entry[5]
            log_aggregator[key][2] += entry[6]
        else:
            log_with_trailer.append(entry + (None,))

for (assembly, material, group, trailer), (qty, wt, vol) in log_aggregator.items():
    log_with_trailer.append((assembly, material, group, "Loaded", qty, wt, vol, trailer))

# --- Identify assemblies split across multiple trailers and add split trailer info ---
# Track which trailers each (assembly, material) pair appears in
assembly_material_trailers = defaultdict(set)
for row in log_with_trailer:
    if row[3] == "Loaded":
        assembly_material_trailers[(row[0], row[1])].add(row[7])

# Prepare output dataframe and add split trailer column
output_data = pd.DataFrame(log_with_trailer, columns=['Assembly', 'Material', 'Group', 'Status', 'Quantity', 'Weight', 'Volume', 'Trailer'])

def split_trailer_note(row):
    if row['Status'] != "Loaded":
        return None
    trailers = assembly_material_trailers[(row['Assembly'], row['Material'])]
    if len(trailers) > 1:
        other_trailers = sorted(t for t in trailers if t != row['Trailer'])
        return ", ".join(other_trailers)
    return None

output_data['Split Assembly Trailers'] = output_data.apply(split_trailer_note, axis=1)

# Prepare summary dataframe with each trailer separately
used_trailers_summary = []
for tname in best_config:
    specs = trailers[tname]
    state = best_config_state[tname]
    total_wt = state['total_weight']
    total_vol = state['total_volume']
    total_ht = state['total_height']
    weight_pct = round((total_wt / specs['max_weight']) * 100, 2) if total_wt is not None else None
    volume_pct = round((total_vol / specs['max_volume']) * 100, 2) if total_vol is not None else None
    height_pct = round((total_ht / specs['height']) * 100, 2) if total_ht is not None else None
    used_trailers_summary.append({
        'Trailer': tname,
        'Total Weight Used (lbs)': total_wt,
        'Total Volume Used (ft³)': total_vol,
        'Weight Used (%)': f"{weight_pct}%" if weight_pct is not None else None,
        'Volume Used (%)': f"{volume_pct}%" if volume_pct is not None else None,
        'Height Used (%)': f"{height_pct}%" if height_pct is not None else None,
        'Max Weight (lbs)': specs['max_weight'],
        'Max Volume (ft³)': specs['max_volume'],
        'Height Used (in)': total_ht,
        'Dimensions (LxWxH in)': f"{specs['length']}x{specs['width']}x{specs['height']}",
        'Total Cost ($)': cost_map[tname],
        'Note': 'Selected Trailer'
    })

# Create unloaded items dataframe
unloaded_df = pd.DataFrame(best_config_unloaded, columns=['Assembly', 'Material', 'Group', 'Reason', 'Quantity', 'Weight', 'Volume', 'Length', 'Width'])

# Output results
summary_df = pd.DataFrame(used_trailers_summary)

with pd.ExcelWriter(output_file, engine='openpyxl') as writer:
    output_data.to_excel(writer, sheet_name='Loading Plan', index=False)
    summary_df.to_excel(writer, sheet_name='Summary', index=False)
    unloaded_df.to_excel(writer, sheet_name='Unloaded Items', index=False)

    # Stack Breakdown Sheet
    stack_rows = []
    for tname in best_config:
        specs = trailers[tname]
        stack_heights = best_config_state[tname].get('stack_heights', [])
        for i, height in enumerate(stack_heights, start=1):
            stack_rows.append({
                'Trailer': tname,
                'Stack Number': i,
                'Stack Height (in)': round(height, 2),
                'Stack Height Used (%)': f"{round((height/specs['height'])*100, 2)}%"
            })
    stack_df = pd.DataFrame(stack_rows)
    stack_df.to_excel(writer, sheet_name='Stack Breakdown', index=False)

# --- Apply formatting and auto-width ---
import openpyxl
from openpyxl.utils import get_column_letter

# Reopen the workbook after writing
wb = openpyxl.load_workbook(output_file)

# Formatting maps for specific columns
def auto_width(ws):
    for col in ws.columns:
        max_length = 0
        column = col[0].column_letter  # Get the column name
        for cell in col:
            try:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
            except:
                pass
        adjusted_width = max_length + 2
        ws.column_dimensions[column].width = adjusted_width

# Format for Loading Plan sheet
ws1 = wb['Loading Plan']
col_map = {cell.value: idx+1 for idx, cell in enumerate(ws1[1])}
for row in ws1.iter_rows(min_row=2, max_row=ws1.max_row):
    # Weight (commas, 2 decimals)
    if 'Weight' in col_map:
        cell = row[col_map['Weight']-1]
        if isinstance(cell.value, (int, float)):
            cell.number_format = '#,##0.00'
    # Volume (commas, 2 decimals)
    if 'Volume' in col_map:
        cell = row[col_map['Volume']-1]
        if isinstance(cell.value, (int, float)):
            cell.number_format = '#,##0.00'
    # Quantity: keep as general
    if 'Quantity' in col_map:
        cell = row[col_map['Quantity']-1]
        cell.number_format = 'General'
    # Trailer: text
    if 'Trailer' in col_map:
        cell = row[col_map['Trailer']-1]
        cell.number_format = '@'

auto_width(ws1)

# Format for Summary sheet
ws2 = wb['Summary']
col_map2 = {cell.value: idx+1 for idx, cell in enumerate(ws2[1])}
for row in ws2.iter_rows(min_row=2, max_row=ws2.max_row):
    # Total Weight Used (commas, 2 decimals)
    if 'Total Weight Used (lbs)' in col_map2:
        cell = row[col_map2['Total Weight Used (lbs)']-1]
        if isinstance(cell.value, (int, float)):
            cell.number_format = '#,##0.00'
    # Total Volume Used (commas, 2 decimals)
    if 'Total Volume Used (ft³)' in col_map2:
        cell = row[col_map2['Total Volume Used (ft³)']-1]
        if isinstance(cell.value, (int, float)):
            cell.number_format = '#,##0.00'
    # Weight Used (%) and Volume Used (%) and Height Used (%): keep as string
    # Max Weight/Volume: commas, no decimals
    if 'Max Weight (lbs)' in col_map2:
        cell = row[col_map2['Max Weight (lbs)']-1]
        if isinstance(cell.value, (int, float)):
            cell.number_format = '#,##0'
    if 'Max Volume (ft³)' in col_map2:
        cell = row[col_map2['Max Volume (ft³)']-1]
        if isinstance(cell.value, (int, float)):
            cell.number_format = '#,##0'
    # Height Used (in): General
    if 'Height Used (in)' in col_map2:
        cell = row[col_map2['Height Used (in)']-1]
        cell.number_format = 'General'
    # Dimensions: General text
    if 'Dimensions (LxWxH in)' in col_map2:
        cell = row[col_map2['Dimensions (LxWxH in)']-1]
        cell.number_format = '@'
    # Total Cost ($): commas, no decimals
    if 'Total Cost ($)' in col_map2:
        cell = row[col_map2['Total Cost ($)']-1]
        if isinstance(cell.value, (int, float)):
            cell.number_format = '#,##0'

auto_width(ws2)

# Format for Unloaded Items sheet
ws3 = wb['Unloaded Items']
col_map3 = {cell.value: idx+1 for idx, cell in enumerate(ws3[1])}
for row in ws3.iter_rows(min_row=2, max_row=ws3.max_row):
    # Weight (commas, 2 decimals)
    if 'Weight' in col_map3:
        cell = row[col_map3['Weight']-1]
        if isinstance(cell.value, (int, float)):
            cell.number_format = '#,##0.00'
    # Volume (commas, 2 decimals)
    if 'Volume' in col_map3:
        cell = row[col_map3['Volume']-1]
        if isinstance(cell.value, (int, float)):
            cell.number_format = '#,##0.00'
    # Quantity: keep as general
    if 'Quantity' in col_map3:
        cell = row[col_map3['Quantity']-1]
        cell.number_format = 'General'
    # Reason: general text
    if 'Reason' in col_map3:
        cell = row[col_map3['Reason']-1]
        cell.number_format = '@'

auto_width(ws3)

# Format for Stack Breakdown sheet
ws4 = wb['Stack Breakdown']
auto_width(ws4)


wb.save(output_file)

# --- Visualization of trailer layout ---
import matplotlib.pyplot as plt
import matplotlib.patches as patches
import os

def generate_trailer_layout_visuals(trailers_state, output_dir="visuals"):
    os.makedirs(output_dir, exist_ok=True)

    for tname, state in trailers_state.items():
        specs = trailers[tname]
        trailer_length = specs['length']
        trailer_width = specs['width']

        fig, ax = plt.subplots(figsize=(12, 6))
        ax.set_xlim(0, trailer_length)
        ax.set_ylim(0, trailer_width)
        ax.set_title(f"Trailer Layout: {tname}")
        ax.set_xlabel("Length (in)")
        ax.set_ylabel("Width (in)")
        ax.set_aspect('equal')
        ax.grid(True, linestyle="--", alpha=0.3)

        # Use color map for assemblies
        # color_map = {}
        # next_color = plt.cm.get_cmap("tab20")

        x_offset = 0
        y_offset = 0

        stack_heights = state.get('stack_heights', [])
        stack_heights_iter = iter(stack_heights)

        # Find unique stacks to display, only for successfully packed physical stacks
        # We'll align stack rectangles to stack_heights (one per height).
        # For labeling, collect loaded log entries in order, one per stack.
        loaded_logs = [log for log in state['log'] if log[3] == "Loaded"]
        # Only as many as stack_heights
        stacks_to_display = []
        for i in range(min(len(stack_heights), len(loaded_logs))):
            stacks_to_display.append(loaded_logs[i])

        for stack_idx, height in enumerate(stack_heights):
            # If we have a log entry for this stack, use its info; else use dummy info
            if stack_idx < len(stacks_to_display):
                log = stacks_to_display[stack_idx]
                assembly = log[0]
                material = log[1]
                qty = log[4]
                length = log[7]
                width = log[8]
            else:
                assembly = ""
                material = ""
                qty = ""
                length = sheet_length
                width = sheet_width
            group = log[2] if stack_idx < len(stacks_to_display) else 'unknown'
            if group == 'wall':
                base_color = 'Blues'
            elif group == 'roof':
                base_color = 'Oranges'
            else:
                base_color = 'Greys'
            cmap = plt.get_cmap(base_color)
            color = cmap(min(1.0, height / specs['height']))
            rect = patches.Rectangle(
                (x_offset, y_offset),
                length,
                width,
                linewidth=1,
                edgecolor='black',
                facecolor=color,
                alpha=0.9
            )
            ax.add_patch(rect)
            ax.text(
                x_offset + length / 2,
                y_offset + width / 2,
                f"{'Stack ' + str(stack_idx+1)}\n{height:.2f} in",
                ha='center', va='center', fontsize=6
            )

            y_offset += width + gap_between_stacks
            if y_offset + width > trailer_width:
                y_offset = 0
                x_offset += length + gap_between_stacks_length
            if x_offset + length > trailer_length:
                break  # no more room

        # Save figure
        plt.tight_layout()
        plt.savefig(os.path.join(output_dir, f"{tname.replace(' ', '_')}_layout.png"))
        plt.close()

# --- Interactive 2D Visualization using Plotly ---
def generate_interactive_2d_layouts(trailers_state, output_dir="interactive_visuals_2d"):
    import os
    os.makedirs(output_dir, exist_ok=True)

    for tname, state in trailers_state.items():
        specs = trailers[tname]
        trailer_length = specs['length']
        trailer_width = specs['width']

        fig = go.Figure()
        x_offset = 0
        y_offset = 0

        stack_data = [
            {
                "assembly": log[0],
                "material": log[1],
                "group": log[2],
                "length": log[7],
                "width": log[8],
                "height": height
            }
            for log, height in zip(state["log"], state.get("stack_heights", []))
            if log[3] == "Loaded"
        ]

        for stack in stack_data:
            assembly = stack["assembly"]
            material = stack["material"]
            group = stack["group"]
            length = stack["length"]
            width = stack["width"]
            height = stack["height"]
            hover_text = f"Assembly: {assembly}<br>Material: {material}<br>Group: {group}<br>Height: {height:.2f} in"

            fig.add_shape(
                type="rect",
                x0=x_offset,
                y0=y_offset,
                x1=x_offset + length,
                y1=y_offset + width,
                line=dict(color="black"),
                fillcolor="blue" if group == "wall" else "orange" if group == "roof" else "gray",
                opacity=0.6,
            )

            fig.add_trace(go.Scatter(
                x=[x_offset + length / 2],
                y=[y_offset + width / 2],
                text=[hover_text],
                mode="text",
                hoverinfo="text"
            ))

            y_offset += width + gap_between_stacks
            if y_offset + width > trailer_width:
                y_offset = 0
                x_offset += length + gap_between_stacks_length
            if x_offset + length > trailer_length:
                break

        fig.update_layout(
            title=f"Interactive 2D Trailer Layout: {tname}",
            xaxis=dict(range=[0, trailer_length], title="Length (in)", scaleanchor="y", scaleratio=1),
            yaxis=dict(range=[0, trailer_width], title="Width (in)"),
            width=1000,
            height=500,
            showlegend=False
        )

        fig.write_html(os.path.join(output_dir, f"{tname.replace(' ', '_')}_interactive_2d.html"))

# Call visualization after Excel file is saved
generate_trailer_layout_visuals(best_config_state)

# --- Interactive 3D Visualization using plotly ---
def generate_interactive_3d_layouts(trailers_state, output_dir="interactive_visuals_3d"):
    import os
    import plotly.graph_objects as go
    os.makedirs(output_dir, exist_ok=True)

    def create_box(x, y, z, dx, dy, dz, group, hovertext, intensity, colorscale):
        vertices = [
            [x, y, z],
            [x+dx, y, z],
            [x+dx, y+dy, z],
            [x, y+dy, z],
            [x, y, z+dz],
            [x+dx, y, z+dz],
            [x+dx, y+dy, z+dz],
            [x, y+dy, z+dz],
        ]
        faces = [
            [0,1,2], [0,2,3],
            [4,5,6], [4,6,7],
            [0,1,5], [0,5,4],
            [1,2,6], [1,6,5],
            [2,3,7], [2,7,6],
            [3,0,4], [3,4,7],
        ]
        x_vals, y_vals, z_vals = zip(*vertices)
        i, j, k = zip(*faces)
        mesh = go.Mesh3d(
            x=x_vals, y=y_vals, z=z_vals,
            i=i, j=j, k=k,
            intensity=[intensity]*len(x_vals),
            colorscale=colorscale,
            cmin=0,
            cmax=1,
            opacity=0.85,
            hovertext=hovertext,
            hoverinfo="text",
            showscale=False,
            lighting=dict(ambient=0.5, diffuse=0.8, roughness=0.9, specular=0.2),
            flatshading=True
        )
        # Edge overlay for visual distinction between stacks/groups
        edge_x = []
        edge_y = []
        edge_z = []
        # 12 edges of a cuboid
        edges = [
            (0,1), (1,2), (2,3), (3,0), # bottom
            (4,5), (5,6), (6,7), (7,4), # top
            (0,4), (1,5), (2,6), (3,7)
        ]
        for a, b in edges:
            edge_x += [vertices[a][0], vertices[b][0], None]
            edge_y += [vertices[a][1], vertices[b][1], None]
            edge_z += [vertices[a][2], vertices[b][2], None]
        edge_trace = go.Scatter3d(
            x=edge_x, y=edge_y, z=edge_z,
            mode='lines',
            line=dict(color='black', width=2),
            hoverinfo='skip',
            showlegend=False,
            opacity=0.5
        )
        return [mesh, edge_trace]

    for tname, state in trailers_state.items():
        specs = trailers[tname]
        trailer_length = specs['length']
        trailer_width = specs['width']
        trailer_height = specs['height']

        fig = go.Figure()
        x_offset = 0
        y_offset = 0
        z_offset = 0

        stack_data = [
            {
                "assembly": log[0],
                "material": log[1],
                "group": log[2],
                "length": log[7],
                "width": log[8],
                "height": height
            }
            for log, height in zip(state["log"], state.get("stack_heights", []))
            if log[3] == "Loaded"
        ]

        # Track last stack position for group separation
        last_x, last_y, last_z = 0, 0, 0
        last_group = None
        stack_z_offsets = []
        # For each stack, determine z_offset based on stacking and group
        x_offset = 0
        y_offset = 0
        z_offset = 0
        prev_stack = None
        for i, stack in enumerate(stack_data):
            group = stack["group"].lower()
            height = stack["height"]
            # --- Color and intensity logic ---
            if group == 'wall':
                colorscale = [[0, 'rgb(173, 216, 230)'], [1, 'rgb(0, 0, 255)']]
                intensity = min(height / 100, 1)
            elif group == 'roof':
                colorscale = [[0, 'rgb(255, 200, 100)'], [1, 'rgb(255, 140, 0)']]
                intensity = min(height / 100, 1)
            else:
                colorscale = [[0, 'gray'], [1, 'black']]
                intensity = 0.5

            # Determine z_offset for stacking and group separation
            if i == 0:
                z_offset = 0
            else:
                prev = stack_data[i-1]
                prev_group = prev["group"].lower()
                prev_height = prev["height"]
                if (abs(x_offset - last_x) < 1e-6 and abs(y_offset - last_y) < 1e-6):
                    # If same stack position, stack on top
                    if group != prev_group:
                        # If group changes at same stack position, offset a bit for visibility
                        z_offset += prev_height * 0.98
                        # Visual separation: slightly different z to avoid overlap
                        # Optionally, change opacity or color
                        intensity = 1.0 if group == 'roof' else 0.6
                    else:
                        z_offset += prev_height
                else:
                    z_offset = 0
            last_x, last_y, last_z = x_offset, y_offset, z_offset
            # --- End stacking logic ---

            assembly = stack["assembly"]
            material = stack["material"]
            length = stack["length"]
            width = stack["width"]
            hover_text = (
                f"Assembly: {assembly}<br>Material: {material}<br>"
                f"Group: {stack['group']}<br>Height: {height:.2f} in"
            )
            traces = create_box(
                x_offset, y_offset, z_offset,
                length, width, height,
                stack["group"],
                hover_text,
                intensity,
                colorscale
            )
            for trace in traces:
                fig.add_trace(trace)

            # Update for next stack
            y_offset += width + gap_between_stacks
            if y_offset + width > trailer_width:
                y_offset = 0
                x_offset += length + gap_between_stacks_length
            if x_offset + length > trailer_length:
                break

        fig.update_layout(
            title=f"Interactive 3D Trailer Layout: {tname}",
            scene=dict(
                xaxis_title="Length (in)",
                yaxis_title="Width (in)",
                zaxis_title="Height (in)",
                xaxis=dict(range=[0, trailer_length]),
                yaxis=dict(range=[0, trailer_width]),
                zaxis=dict(range=[0, trailer_height])
            ),
            width=1000,
            height=700,
            margin=dict(l=0, r=0, t=40, b=0)
        )

        fig.write_html(os.path.join(output_dir, f"{tname.replace(' ', '_')}_interactive_3d.html"))

# Call interactive 3D visualization
generate_interactive_3d_layouts(best_config_state)

# --- Interactive 2D and 3D Visualizations ---

# Interactive 2D layout function (if not already present)
def generate_interactive_2d_layouts(trailers_state, output_dir="interactive_visuals_2d"):
    import plotly.graph_objects as go
    import os
    os.makedirs(output_dir, exist_ok=True)
    for tname, state in trailers_state.items():
        specs = trailers[tname]
        trailer_length = specs['length']
        trailer_width = specs['width']
        fig = go.Figure()
        x_offset = 0
        y_offset = 0
        stack_data = [
            {
                "assembly": log[0],
                "material": log[1],
                "group": log[2],
                "length": log[7],
                "width": log[8],
                "height": height
            }
            for log, height in zip(state["log"], state.get("stack_heights", []))
            if log[3] == "Loaded"
        ]
        for stack in stack_data:
            assembly = stack["assembly"]
            material = stack["material"]
            group = stack["group"]
            length = stack["length"]
            width = stack["width"]
            height = stack["height"]
            hover_text = f"Assembly: {assembly}<br>Material: {material}<br>Group: {group}<br>Height: {height:.2f} in"
            color = "blue" if group == "wall" else "orange" if group == "roof" else "gray"
            fig.add_shape(
                type="rect",
                x0=x_offset, y0=y_offset,
                x1=x_offset+length, y1=y_offset+width,
                line=dict(color="black"),
                fillcolor=color,
                opacity=0.7,
            )
            fig.add_trace(go.Scatter(
                x=[x_offset + length/2],
                y=[y_offset + width/2],
                text=[hover_text],
                mode="text",
                showlegend=False,
                hoverinfo="text"
            ))
            y_offset += width + gap_between_stacks
            if y_offset + width > trailer_width:
                y_offset = 0
                x_offset += length + gap_between_stacks_length
            if x_offset + length > trailer_length:
                break
        fig.update_layout(
            title=f"Interactive 2D Trailer Layout: {tname}",
            xaxis=dict(range=[0, trailer_length], title="Length (in)", constrain='domain', scaleanchor='y', scaleratio=1),
            yaxis=dict(range=[0, trailer_width], title="Width (in)", constrain='domain'),
            width=900, height=500, margin=dict(l=0, r=0, t=40, b=0),
            plot_bgcolor="white",
        )
        fig.write_html(os.path.join(output_dir, f"{tname.replace(' ', '_')}_interactive_2d.html"))

def generate_interactive_3d_layouts(trailers_state, output_dir="interactive_visuals_3d"):
    import os
    import plotly.graph_objects as go
    os.makedirs(output_dir, exist_ok=True)

    def create_box(x, y, z, dx, dy, dz, group, hovertext):
        vertices = [
            [x, y, z],
            [x+dx, y, z],
            [x+dx, y+dy, z],
            [x, y+dy, z],
            [x, y, z+dz],
            [x+dx, y, z+dz],
            [x+dx, y+dy, z+dz],
            [x, y+dy, z+dz],
        ]
        faces = [
            [0,1,2], [0,2,3],
            [4,5,6], [4,6,7],
            [0,1,5], [0,5,4],
            [1,2,6], [1,6,5],
            [2,3,7], [2,7,6],
            [3,0,4], [3,4,7],
        ]
        x_vals, y_vals, z_vals = zip(*vertices)
        i, j, k = zip(*faces)

        if group == "wall":
            intensity = 0.2
            colorscale = [[0, 'rgb(173, 216, 230)'], [1, 'rgb(0, 0, 255)']]
        elif group == "roof":
            intensity = 0.8
            colorscale = [[0, 'rgb(255, 200, 0)'], [1, 'rgb(255, 100, 0)']]
        else:
            intensity = 0.5
            colorscale = [[0, 'rgb(200, 200, 200)'], [1, 'rgb(100, 100, 100)']]

        return go.Mesh3d(
            x=x_vals, y=y_vals, z=z_vals,
            i=i, j=j, k=k,
            intensity=[intensity]*len(x_vals),
            colorscale=colorscale,
            cmin=0,
            cmax=1,
            opacity=0.75,
            hovertext=hovertext,
            hoverinfo="text",
            showscale=False,
            lighting=dict(ambient=0.5, diffuse=0.8, roughness=0.9, specular=0.2),
            flatshading=True
        )

    for tname, state in trailers_state.items():
        specs = trailers[tname]
        trailer_length = specs['length']
        trailer_width = specs['width']
        trailer_height = specs['height']

        fig = go.Figure()
        x_offset = 0
        y_offset = 0

        stack_data = [
            {
                "assembly": log[0],
                "material": log[1],
                "group": log[2],
                "length": log[7],
                "width": log[8],
                "height": height
            }
            for log, height in zip(state["log"], state.get("stack_heights", []))
            if log[3] == "Loaded"
        ]

        for stack in stack_data:
            assembly = stack["assembly"]
            material = stack["material"]
            group = stack["group"]
            length = stack["length"]
            width = stack["width"]
            height = stack["height"]
            hover_text = f"Assembly: {assembly}<br>Material: {material}<br>Group: {group}<br>Height: {height:.2f} in"

            traces = create_box(
                x_offset, y_offset, 0,
                length, width, height,
                group,
                hover_text
            )
            # Ensure traces is a list/iterable of traces, or a single trace
            if not isinstance(traces, (list, tuple)):
                traces = [traces]
            for trace in traces:
                print("TRACE TYPE:", type(trace))
                if isinstance(trace, (go.Mesh3d, go.Scatter3d)):
                    fig.add_trace(trace)

            y_offset += width + gap_between_stacks
            if y_offset + width > trailer_width:
                y_offset = 0
                x_offset += length + gap_between_stacks_length
            if x_offset + length > trailer_length:
                break

        fig.update_layout(
            title=f"Trailer: {tname} - Interactive 3D Layout",
            scene=dict(
                xaxis=dict(title='Length', range=[0, specs['length']]),
                yaxis=dict(title='Width', range=[0, specs['width']]),
                zaxis=dict(title='Height', range=[0, specs['height']])
            ),
            margin=dict(l=0, r=0, b=0, t=40)
        )

        fig.write_html(os.path.join(output_dir, f"{tname.replace(' ', '_')}_interactive_3d.html"))

# Call interactive 2D and 3D visualizations after static visualizations
generate_interactive_2d_layouts(best_config_state)
generate_interactive_3d_layouts(best_config_state)

# Call interactive 2D visualization
generate_interactive_2d_layouts(best_config_state)