
import pandas as pd
import openpyxl
from openpyxl.styles import PatternFill
from openpyxl.styles import Border, Side, Alignment, Font
import itertools

# Material density in lb/ft³
MATERIAL_DENSITY = {
    "steel": 490  # lb/ft³
}

# Trailer costs
TRAILER_COSTS = {
    "Tandem": 3000,
    "Flat Deck 53'": 4000,
    "Super B": 5000
}

# Trailer specifications
TRAILER_SPECS = {
    "Tandem": {"max_volume": 3480.32, "max_weight": 47000},
    "Flat Deck 53'": {"max_volume": 3842.85, "max_weight": 46000},
    "Super B": {"max_volume": 4350.39, "max_weight": 92000}
}

PREFERRED_TRAILER = ""  # e.g., "Flat Deck 53'"
SINGLE_TRAILER_ONLY = ""  # Set to a trailer name like "Tandem" to only use that trailer
country = "USA".strip().lower() #input("Enter destination country: ").strip().lower()

# Define output path early for use in conditional notes
output_path = "/Users/craignyabvure/Desktop/Python Test Projects AGI/Trailer Loading/Loading_Plan_Output.xlsx"

# Load Excel data using paths
solver_df = pd.read_excel("/Users/craignyabvure/Desktop/Python Test Projects AGI/Trailer Loading/Solver (2).xlsx")
bom_df = pd.read_excel("/Users/craignyabvure/Desktop/Python Test Projects AGI/Trailer Loading/BOM Parts.xlsx")

# Normalize all column names first
bom_df.columns = bom_df.columns.str.strip().str.lower()
solver_df.columns = solver_df.columns.str.strip().str.lower()

# Rename relevant columns in solver_df
solver_df.rename(columns={'material desc': 'description', 'net weight': 'weight'}, inplace=True)

# Standardize key columns
bom_df.rename(columns={'assembly': 'assembly', 'part#': 'part#', 'quantity': 'quantity'}, inplace=True)
solver_df.rename(columns={'material': 'part#'}, inplace=True)

# Ensure both part# columns are string and lowercase
bom_df['part#'] = bom_df['part#'].astype(str).str.strip().str.lower()
solver_df['part#'] = solver_df['part#'].astype(str).str.strip().str.lower()

# Merge and clean data
merged_df = pd.merge(bom_df, solver_df, how='left', left_on='part#', right_on='part#')
merged_df.dropna(subset=['weight'], inplace=True)
merged_df['material'] = 'steel'  # assuming all parts are steel
merged_df['volume'] = merged_df['weight'] / merged_df['material'].map(MATERIAL_DENSITY)
merged_df['quantity'] = pd.to_numeric(merged_df['quantity'], errors='coerce').fillna(0).astype(int)
merged_df = merged_df[merged_df['quantity'] > 0]

# Remove any volume normalization to keep raw values

# Remove any volume normalization to keep raw values
# Drop duplicates by assembly and part#
merged_df = merged_df.drop_duplicates(subset=['assembly', 'part#'])

restricted = {"america", "us", "usa", "united states", "united states of america"}
trailer_specs_filtered = {
    name: spec for name, spec in TRAILER_SPECS.items()
    if not (name == "Super B" and any(key in country for key in restricted))
}
if SINGLE_TRAILER_ONLY:
    # Restriction: If "Super B" is selected and the country is restricted, output a note and exit.
    if SINGLE_TRAILER_ONLY.lower() == "super b" and any(key in country for key in restricted):
        wb = openpyxl.Workbook()
        ws_note = wb.active
        ws_note.title = "Loading Plan"
        ws_note.append(["Note"])
        ws_note.append(["No output generated. The selected trailer 'Super B' cannot be used for shipments going to or through the USA/America due to restrictions."])
        wb.save(output_path)
        print(f"Loading plan saved to {output_path}")
        exit()
    trailer_specs_filtered = {
        name: spec for name, spec in trailer_specs_filtered.items()
        if name.lower() == SINGLE_TRAILER_ONLY.lower()
    }

output_rows = []
assemblies = merged_df.groupby('assembly')

assembly_groups = list(assemblies)

from itertools import islice

total_required_weight = (merged_df['weight'] * merged_df['quantity']).sum()
total_required_volume = (merged_df['volume'] * merged_df['quantity']).sum()

best_output = []
best_usage = 0

def trailer_sort_key(item):
    name, spec = item
    base_priority = -(spec["max_weight"] * spec["max_volume"])
    if PREFERRED_TRAILER and PREFERRED_TRAILER.lower() == name.lower():
        return float('inf')  # Sort preferred trailer to top
    return base_priority

default_order = sorted(trailer_specs_filtered.items(), key=trailer_sort_key, reverse=True)
alternative_order = default_order[1:] + default_order[:1] if len(default_order) > 1 else default_order
if SINGLE_TRAILER_ONLY:
    alternative_order = default_order

for perm in islice(itertools.permutations(assembly_groups), 200):  # evaluate top 200 permutations
    # Evaluate default order
    trailers_copy = []
    used_weight = 0
    used_volume = 0
    for i, (name, spec) in enumerate(default_order):
        trailers_copy.append({
            "id": i + 1,
            "name": name,
            "max_volume": spec["max_volume"],
            "max_weight": spec["max_weight"] - 5000,
            "used_volume": 0,
            "used_weight": 0,
            "items": []
        })
        used_weight += spec["max_weight"]
        used_volume += spec["max_volume"]

    current_output_rows = []
    for assembly_id, group in perm:
        assigned = False
        group_total_weight = (group['weight'] * group['quantity']).sum()
        group_total_volume = (group['volume'] * group['quantity']).sum()

        for trailer in trailers_copy:
            if (trailer['used_weight'] + group_total_weight <= trailer['max_weight']) and (trailer['used_volume'] + group_total_volume <= trailer['max_volume']):
                for _, row in group.iterrows():
                    trailer['items'].append(row)
                    current_output_rows.append({
                        "Trailer": trailer['name'],
                        "Part ID": row["part#"],
                        "Assembly": row["assembly"],
                        "Description": row.get("description", ""),
                        "Total Weight": row["weight"] * row["quantity"],
                        "Total Volume": row["volume"] * row["quantity"],
                        "Quantity": row["quantity"],
                        "Reason": f"Complete Assembly in {trailer['name']}"
                    })
                trailer["used_weight"] += group_total_weight
                trailer["used_volume"] += group_total_volume
                print(f"{trailer['name']} used volume: {trailer['used_volume']:.3f} m³ / {trailer['max_volume']} m³")
                assigned = True
                break

        if not assigned:
            for _, row in group.iterrows():
                part_weight = row["weight"] * row["quantity"]
                part_volume = row["volume"] * row["quantity"]
                part_assigned = False
                for trailer in trailers_copy:
                    if (trailer['used_weight'] + part_weight <= trailer['max_weight']) and (trailer['used_volume'] + part_volume <= trailer['max_volume']):
                        trailer['items'].append(row)
                        trailer["used_weight"] += part_weight
                        trailer["used_volume"] += part_volume
                        print(f"{trailer['name']} used volume: {trailer['used_volume']:.3f} m³ / {trailer['max_volume']} m³")
                        current_output_rows.append({
                            "Trailer": trailer['name'],
                            "Part ID": row["part#"],
                            "Assembly": row["assembly"],
                            "Description": row.get("description", ""),
                            "Total Weight": part_weight,
                            "Total Volume": part_volume,
                            "Quantity": row["quantity"],
                            "Reason": f"Split from Assembly {row['assembly']}"
                        })
                        part_assigned = True
                        break
                if not part_assigned:
                    current_output_rows.append({
                        "Trailer": "UNASSIGNED",
                        "Part ID": row["part#"],
                        "Assembly": row["assembly"],
                        "Description": row.get("description", ""),
                        "Total Weight": part_weight,
                        "Total Volume": part_volume,
                        "Quantity": row["quantity"],
                        "Reason": "Does not fit in any trailer"
                    })
    usage = sum(trailer["used_weight"] + trailer["used_volume"] for trailer in trailers_copy)
    if all(trailer["used_volume"] <= trailer["max_volume"] and trailer["used_weight"] <= trailer["max_weight"] for trailer in trailers_copy):
        if usage > best_usage:
            best_usage = usage
            best_output = current_output_rows

    # Evaluate alternative order
    alt_trailers_copy = []
    alt_used_weight = 0
    alt_used_volume = 0
    for i, (name, spec) in enumerate(alternative_order):
        alt_trailers_copy.append({
            "id": i + 1,
            "name": name,
            "max_volume": spec["max_volume"],
            "max_weight": spec["max_weight"] - 5000,
            "used_volume": 0,
            "used_weight": 0,
            "items": []
        })
        alt_used_weight += spec["max_weight"]
        alt_used_volume += spec["max_volume"]

    alternative_output_rows = []
    for assembly_id, group in perm:
        assigned = False
        group_total_weight = (group['weight'] * group['quantity']).sum()
        group_total_volume = (group['volume'] * group['quantity']).sum()

        for trailer in alt_trailers_copy:
            if (trailer['used_weight'] + group_total_weight <= trailer['max_weight']) and (trailer['used_volume'] + group_total_volume <= trailer['max_volume']):
                for _, row in group.iterrows():
                    trailer['items'].append(row)
                    alternative_output_rows.append({
                        "Trailer": trailer['name'],
                        "Part ID": row["part#"],
                        "Assembly": row["assembly"],
                        "Description": row.get("description", ""),
                        "Total Weight": row["weight"] * row["quantity"],
                        "Total Volume": row["volume"] * row["quantity"],
                        "Quantity": row["quantity"],
                        "Reason": f"Complete Assembly in {trailer['name']}"
                    })
                trailer["used_weight"] += group_total_weight
                trailer["used_volume"] += group_total_volume
                assigned = True
                break

        if not assigned:
            for _, row in group.iterrows():
                part_weight = row["weight"] * row["quantity"]
                part_volume = row["volume"] * row["quantity"]
                part_assigned = False
                for trailer in alt_trailers_copy:
                    if (trailer['used_weight'] + part_weight <= trailer['max_weight']) and (trailer['used_volume'] + part_volume <= trailer['max_volume']):
                        trailer['items'].append(row)
                        trailer["used_weight"] += part_weight
                        trailer["used_volume"] += part_volume
                        alternative_output_rows.append({
                            "Trailer": trailer['name'],
                            "Part ID": row["part#"],
                            "Assembly": row["assembly"],
                            "Description": row.get("description", ""),
                            "Total Weight": part_weight,
                            "Total Volume": part_volume,
                            "Quantity": row["quantity"],
                            "Reason": f"Split from Assembly {row['assembly']}"
                        })
                        part_assigned = True
                        break
                if not part_assigned:
                    alternative_output_rows.append({
                        "Trailer": "UNASSIGNED",
                        "Part ID": row["part#"],
                        "Assembly": row["assembly"],
                        "Description": row.get("description", ""),
                        "Total Weight": part_weight,
                        "Total Volume": part_volume,
                        "Quantity": row["quantity"],
                        "Reason": "Does not fit in any trailer"
                    })
    alt_usage = sum(trailer["used_weight"] + trailer["used_volume"] for trailer in alt_trailers_copy)
    if all(trailer["used_volume"] <= trailer["max_volume"] and trailer["used_weight"] <= trailer["max_weight"] for trailer in alt_trailers_copy):
        if alt_usage > best_usage and alternative_output_rows != best_output:
            best_usage = alt_usage
            best_output = current_output_rows
            output_rows = current_output_rows
            alternative_output_rows_final = alternative_output_rows
        else:
            alternative_output_rows_final = alternative_output_rows
    else:
        alternative_output_rows_final = alternative_output_rows

output_rows = best_output

wb = openpyxl.Workbook()
ws = wb.active
ws.title = "Loading Plan"

headers = ["Trailer", "Assembly", "Part ID", "Description", "Total Weight (lbs)", "Total Volume (ft³)", "Quantity", "Reason"]
ws.append(headers)

fills = {
    "Tandem": PatternFill(start_color="ADD8E6", end_color="ADD8E6", fill_type="solid"),
    "Flat Deck 53'": PatternFill(start_color="90EE90", end_color="90EE90", fill_type="solid"),
    "Super B": PatternFill(start_color="FFFF99", end_color="FFFF99", fill_type="solid"),
    "UNASSIGNED": PatternFill(start_color="FF9999", end_color="FF9999", fill_type="solid"),
}

for row_data in output_rows:
    # Provide fallback for any missing keys for backward compatibility
    row = [
        row_data.get("Trailer", ""),
        row_data.get("Assembly", ""),
        row_data.get("Part ID", ""),
        row_data.get("Description", ""),
        row_data.get("Total Weight", ""),
        row_data.get("Total Volume", ""),
        row_data.get("Quantity", ""),
        row_data.get("Reason", "")
    ]
    ws.append(row)
    trailer = row_data["Trailer"]
    if trailer in fills:
        for cell in ws[ws.max_row]:
            cell.fill = fills[trailer]


# Create second sheet for trailer summary
ws_summary = wb.create_sheet(title="Trailer Summary")
ws_summary.append(["Truck", "Total Quantity", "Total Weight", "Total Volume", "Weight % Used", "Volume % Used", "Max Weight", "Max Volume"])
trailer_summary = {}

for row in output_rows:
    trailer = row["Trailer"]
    if trailer == "UNASSIGNED":
        continue
    if trailer not in trailer_summary:
        # Use the global trailer specs dict
        trailer_lookup = TRAILER_SPECS
        trailer_summary[trailer] = {
            "qty": 0,
            "weight": 0,
            "volume": 0,
            "max_weight": trailer_lookup[trailer]["max_weight"],
            "max_volume": trailer_lookup[trailer]["max_volume"]
        }
    trailer_summary[trailer]["qty"] += row["Quantity"]
    trailer_summary[trailer]["weight"] += row["Total Weight"]
    trailer_summary[trailer]["volume"] += row["Total Volume"]

for trailer_name, data in trailer_summary.items():
    weight_pct = (data["weight"] / data["max_weight"]) * 100 if data["max_weight"] else 0
    volume_pct = (data["volume"] / data["max_volume"]) * 100 if data["max_volume"] else 0
    ws_summary.append([
        trailer_name,
        data["qty"],
        data["weight"],
        data["volume"],
        f"{weight_pct:.2f}%",
        f"{volume_pct:.2f}%",
        data["max_weight"],
        data["max_volume"]
    ])

# Add total cost for default method
total_cost = sum(TRAILER_COSTS.get(name, 0) for name in trailer_summary)
ws_summary.append([])
ws_summary.append(["Total Cost", f"${total_cost}"])

ws_summary.append([])
ws_summary.append(["Alternative Sort (Preferred First)"])
ws_summary.append(["Truck", "Total Quantity", "Total Weight", "Total Volume", "Weight % Used", "Volume % Used", "Max Weight", "Max Volume"])
sorted_keys = sorted(trailer_summary.keys(), key=lambda k: (0 if PREFERRED_TRAILER and k.lower() == PREFERRED_TRAILER.lower() else 1, k.lower())) if PREFERRED_TRAILER else sorted(trailer_summary.keys())
alternative_trailer_summary = {}

for row in alternative_output_rows_final:
    trailer = row["Trailer"]
    if trailer == "UNASSIGNED":
        continue
    if trailer not in alternative_trailer_summary:
        trailer_lookup = TRAILER_SPECS
        alternative_trailer_summary[trailer] = {
            "qty": 0,
            "weight": 0,
            "volume": 0,
            "max_weight": trailer_lookup[trailer]["max_weight"],
            "max_volume": trailer_lookup[trailer]["max_volume"]
        }
    alternative_trailer_summary[trailer]["qty"] += row["Quantity"]
    alternative_trailer_summary[trailer]["weight"] += row["Total Weight"]
    alternative_trailer_summary[trailer]["volume"] += row["Total Volume"]

for trailer_name in sorted_keys:
    if trailer_name not in alternative_trailer_summary:
        continue
    data = alternative_trailer_summary[trailer_name]
    weight_pct = (data["weight"] / data["max_weight"]) * 100 if data["max_weight"] else 0
    volume_pct = (data["volume"] / data["max_volume"]) * 100 if data["max_volume"] else 0
    ws_summary.append([
        trailer_name,
        data["qty"],
        data["weight"],
        data["volume"],
        f"{weight_pct:.2f}%",
        f"{volume_pct:.2f}%",
        data["max_weight"],
        data["max_volume"]
    ])

# Add alternative total cost
alt_total_cost = sum(TRAILER_COSTS.get(name, 0) for name in alternative_trailer_summary)
ws_summary.append([])
ws_summary.append(["Alternative Total Cost", f"${alt_total_cost}"])

# Create third sheet for assembly status
ws_assembly = wb.create_sheet(title="Assembly Fit Summary")
ws_assembly.append(["Assembly", "Total Quantity", "Fitted Quantity", "Unfitted Quantity"])
assembly_status = {}


# First pass: collect total quantity per assembly
for row in output_rows:
    asm = row["Assembly"]
    qty = row["Quantity"]
    if asm not in assembly_status:
        assembly_status[asm] = {"total": 0, "fit": 0, "unfit": 0}
    assembly_status[asm]["total"] += qty

# Second pass: classify quantities as fit or unfit
for row in output_rows:
    asm = row["Assembly"]
    qty = row["Quantity"]
    if row["Trailer"] == "UNASSIGNED":
        assembly_status[asm]["unfit"] += qty
    else:
        assembly_status[asm]["fit"] += qty

for asm, data in assembly_status.items():
    ws_assembly.append([
        asm,
        data["total"],
        data["fit"],
        data["unfit"]
    ])


# Create fourth sheet for Unassigned Parts
ws_unassigned = wb.create_sheet(title="Unassigned Parts")
ws_unassigned.append(["Assembly", "Part ID", "Description", "Quantity", "Total Weight (lbs)", "Total Volume (ft³)", "Assigned Trailer (if partial)"])

# Build a mapping of assemblies to trailers (for partials)
assembly_to_trailers = {}
for row in output_rows:
    if row["Trailer"] != "UNASSIGNED":
        asm = row["Assembly"]
        if asm not in assembly_to_trailers:
            assembly_to_trailers[asm] = set()
        assembly_to_trailers[asm].add(row["Trailer"])

# Append unassigned parts and possible assigned trailer info, sorted by Assembly
for row in sorted(output_rows, key=lambda x: x["Assembly"]):
    if row["Trailer"] == "UNASSIGNED":
        asm = row["Assembly"]
        assigned_trailers = assembly_to_trailers.get(asm, [])
        if assigned_trailers:
            assigned_to = f"Partially Loaded in: {', '.join(sorted(assigned_trailers))}"
        else:
            assigned_to = "Not Loaded"
        ws_unassigned.append([
            asm,
            row["Part ID"],
            row.get("Description", ""),
            row["Quantity"],
            row["Total Weight"],
            row["Total Volume"],
            assigned_to
        ])

# If you want to generate a summary string of unassigned trailer parts (for possible reporting):
unassigned_trailer_info = {}
for row in output_rows:
    if row["Trailer"] == "UNASSIGNED":
        asm = row["Assembly"]
        unassigned_trailer_info[asm] = unassigned_trailer_info.get(asm, 0) + row["Quantity"]
# Example formatting (updated per instructions)
if unassigned_trailer_info:
    trailer_parts = [f"{int(qty)} part{'s' if int(qty) > 1 else ''} in {name}" for name, qty in sorted(unassigned_trailer_info.items())]
    unassigned_parts_str = "; ".join(trailer_parts)
    # print or use unassigned_parts_str as needed


# Formatting function for sheets
def format_sheet(ws):
    thin = Side(border_style="thin", color="000000")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    for row in ws.iter_rows():
        for cell in row:
            cell.border = border
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            if cell.row == 1:
                cell.font = Font(bold=True)
    for col in ws.columns:
        max_length = 0
        col_letter = col[0].column_letter
        for cell in col:
            try:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
            except:
                pass
        ws.column_dimensions[col_letter].width = max_length + 2

# Apply formatting to all relevant sheets before saving
for ws_to_format in [ws, ws_summary, ws_assembly, ws_unassigned]:
    format_sheet(ws_to_format)

wb.save(output_path)
print(f"Loading plan saved to {output_path}")
