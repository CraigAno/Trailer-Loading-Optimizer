import pandas as pd
import math
import itertools
import re
import plotly.graph_objects as go
import tempfile
import os
import matplotlib.pyplot as plt
import matplotlib.patches as patches
from collections import defaultdict


  # Default buffer in inches, can be adjusted as needed

def get_effective_trailer_dimensions(trailer_specs, buffer):
    return {
        'length': trailer_specs['length'] - 2 * buffer,
        'width': trailer_specs['width'] - 2 * buffer,
        'height': trailer_specs['height'],
        'max_volume': trailer_specs['max_volume'],
        'max_weight': trailer_specs['max_weight']
    }

def run_trailer_optimizer(bom_file, solver_file, roof_dims_file, country, buffer):
    # --- Constants and Configuration ---
    # Trailer specs (in inches for dimensions)
    trailers = {
        'Tandem': {'length': 624, 'width': 96, 'height': 96, 'max_volume': 3480.32, 'max_weight': 47000},
        'Flat Deck 53': {'length': 636, 'width': 102, 'height': 96, 'max_volume': 3842.85, 'max_weight': 46000},
        'Super B': {'length': 360, 'width': 102, 'height': 96, 'max_volume': 4350.39, 'max_weight': 92000}
    }
    cost_map = {'Tandem': 5000, 'Flat Deck 53': 6500, 'Super B': 7500}
    restricted = {"america", "us", "usa", "united states", "united states of america"}
    sheet_length = 116.5
    sheet_width = 45.5
    sheet_height = 0.25
    sheet_volume = 0.767  # cubic feet
    gap_between_stacks = 3.5
    gap_between_stacks_length = 2.5
    stack_height_limit_pct = 0.92

    # --- Load Excel Data ---
    bom_df = pd.read_excel(bom_file, engine="openpyxl")
    solver_df = pd.read_excel(solver_file, engine="openpyxl")

    # Normalize column names
    bom_df.columns = bom_df.columns.str.strip().str.lower()
    solver_df.columns = solver_df.columns.str.strip().str.lower()

    # Detect quantity column
    qty_col_candidates = ['quantity', 'qty']
    qty_col = next((col for col in qty_col_candidates if col in bom_df.columns), None)
    if not qty_col:
        raise ValueError("No Quantity or Qty column found in BOM")
    # Set quantity column to numeric
    bom_df['quantity'] = pd.to_numeric(bom_df[qty_col], errors='coerce').fillna(0).astype(int)

    # Group to sum duplicate assembly-material pairs
    bom_df = bom_df.groupby(['assembly', 'material'], as_index=False).agg({'quantity':'sum'})

    # Continue as before
    bom_df.rename(columns={'assembly': 'assembly', 'material': 'material', 'quantity': 'quantity'}, inplace=True)
    solver_df.rename(columns={'material': 'material', 'material desc': 'description', 'matl grou': 'group', 'net weight': 'weight'}, inplace=True)
    bom_df['material'] = bom_df['material'].astype(str).str.strip().str.lower()
    solver_df['material'] = solver_df['material'].astype(str).str.strip().str.lower()

    merged_df = pd.merge(bom_df, solver_df, on='material', how='left')

    def extract_bin(material_code):
        match = re.match(r"(\d+)b", material_code.lower())
        return match.group(1) if match else None

    solver_df['roof_bin'] = solver_df.apply(
        lambda row: extract_bin(row['material']) if row['group'] == 'roof' else None,
        axis=1
    )

    unmatched_bom = bom_df[~bom_df['material'].isin(solver_df['material'])].copy()
    unmatched_bom['group'] = 'unknown'
    unmatched_bom['weight'] = None
    unmatched_bom['reason'] = 'Material not found in solver database'
    unmatched_bom = unmatched_bom[['assembly', 'material', 'group', 'quantity', 'weight', 'reason']]

    merged_df['assembly'] = merged_df['assembly'].astype(str).str.strip().str.lower()
    merged_df['group'] = merged_df['group'].astype(str).str.strip().str.lower()
    merged_df['weight'] = pd.to_numeric(merged_df['weight'], errors='coerce')
    required_columns = ['assembly', 'material', 'quantity', 'group', 'weight']
    for col in required_columns:
        if col not in merged_df.columns:
            raise ValueError(f"Required column '{col}' not found in the merged data.")
    merged_df = merged_df.groupby(['assembly', 'material', 'group', 'weight'], as_index=False)['quantity'].sum()

    # --- ROOF SHEET DIMENSIONS HANDLING ---
    def prepare_roof_items(merged_df, roof_dims_path):
        roof_df = pd.read_excel(roof_dims_path, engine="openpyxl")
        roof_df.columns = roof_df.columns.str.strip().str.lower()
        roof_df['material'] = roof_df['material'].astype(str).str.strip().str.lower()
        roof_items = merged_df[merged_df['group'] == 'roof'].copy()
        roof_items['roof_bin'] = roof_items['material'].apply(lambda mat: extract_bin(mat))
        roof_items = roof_items[roof_items['roof_bin'].notna()]
        roof_items = roof_items.merge(roof_df, left_on='roof_bin', right_on='material', how='left', suffixes=('', '_roofdim'))
        for col in ['length', 'width', 'height']:
            if col not in roof_items.columns:
                raise ValueError(f"Missing '{col}' in Roof Sheet Dimensions file.")
        roof_items['length'] = pd.to_numeric(roof_items['length'], errors='coerce')
        roof_items['width'] = pd.to_numeric(roof_items['width'], errors='coerce')
        roof_items['height'] = pd.to_numeric(roof_items['height'], errors='coerce')
        roof_items['volume'] = (roof_items['length'] * roof_items['width'] * roof_items['height']) / 1728
        roof_items.dropna(subset=['length', 'width', 'height', 'volume', 'weight'], inplace=True)
        return roof_items

    # --- FLOOR SHEET DIMENSIONS HANDLING ---
    def prepare_floor_items(merged_df):
        # Fixed file path as per instructions
        floor_path = '/Users/craignyabvure/Desktop/Python Test Projects AGI/Trailer Loading/Floor Dimensions.xlsx'
        floor_df = pd.read_excel(floor_path, engine="openpyxl")
        # Normalize column names
        floor_df.columns = floor_df.columns.str.strip().str.lower()
        print("DEBUG: Detected columns in Floor Dimensions file:", list(floor_df.columns))
        # Lowercase and strip for material
        floor_df['material'] = floor_df['material'].astype(str).str.strip().str.lower()
        # Show a sample of the raw floor dimension data before unit conversion
        print("DEBUG: Sample raw floor dimension data (first 3 rows):\n", floor_df.head(3))
        # Find relevant columns (allowing for "lenght" typo)
        len_col = next((col for col in floor_df.columns if col.startswith('lenght')), None)
        if not len_col:
            len_col = next((col for col in floor_df.columns if col.startswith('length')), None)
        wid_col = next((col for col in floor_df.columns if col.startswith('width')), None)
        ht_col = next((col for col in floor_df.columns if col.startswith('height')), None)
        vol_col = next((col for col in floor_df.columns if col.startswith('volume')), None)
        if not (len_col and wid_col and ht_col and vol_col):
            raise ValueError("Floor Dimensions.xlsx must have columns for 'Lenght(in)', 'Width(in)', 'Height(in)', 'Volume(in^3)' (case/space insensitive)")
        # Keep length, width, height in inches; volume in cubic inches (no conversion)
        floor_df[len_col] = pd.to_numeric(floor_df[len_col], errors='coerce')
        floor_df[wid_col] = pd.to_numeric(floor_df[wid_col], errors='coerce')
        floor_df[ht_col] = pd.to_numeric(floor_df[ht_col], errors='coerce')
        floor_df[vol_col] = pd.to_numeric(floor_df[vol_col], errors='coerce')/1728  # Convert volume from cubic inches to cubic feet
        # Show a sample of the converted length, width, height, and volume values
        print("DEBUG: Sample converted floor dimensions (first 3 rows):")
        print(floor_df[['material', len_col, wid_col, ht_col, vol_col]].head(3))
        # Merge with merged_df for group == 'aerfloor'
        floor_items = merged_df[merged_df['group'] == 'aerfloor'].copy()
        floor_items['material'] = floor_items['material'].astype(str).str.strip().str.lower()
        floor_items = floor_items.merge(
            floor_df[['material', len_col, wid_col, ht_col, vol_col]],
            on='material', how='left', suffixes=('', '_floordim')
        )
        # Rename columns to match roof_items_df shape: length, width, height, volume
        floor_items = floor_items.rename(
            columns={
                len_col: 'length',
                wid_col: 'width',
                ht_col: 'height',
                vol_col: 'volume'
            }
        )
        # Ensure numeric
        floor_items['length'] = pd.to_numeric(floor_items['length'], errors='coerce')
        floor_items['width'] = pd.to_numeric(floor_items['width'], errors='coerce')
        floor_items['height'] = pd.to_numeric(floor_items['height'], errors='coerce')
        floor_items['volume'] = pd.to_numeric(floor_items['volume'], errors='coerce')
        # Show a sample of the merged floor items showing the final dimensions used
        print("DEBUG: Sample merged floor items (first 3 rows):")
        print(floor_items[['assembly', 'material', 'length', 'width', 'height', 'volume', 'weight']].head(3))
        floor_items.dropna(subset=['length', 'width', 'height', 'volume', 'weight'], inplace=True)
        return floor_items

    roof_items_df = prepare_roof_items(merged_df, roof_dims_file)
    floor_items_df = prepare_floor_items(merged_df)
    loadable_groups = ['wall']
    wall_items_df = merged_df[merged_df['group'].isin(loadable_groups)].copy()
    wall_items_df['weight'] = pd.to_numeric(wall_items_df['weight'], errors='coerce')
    wall_items_df['quantity'] = pd.to_numeric(wall_items_df['quantity'], errors='coerce').fillna(0).astype(int)
    # Include wall, roof, and floor items in loadable items
    all_loadable_items_df = pd.concat([wall_items_df, roof_items_df, floor_items_df], ignore_index=True)
    all_loadable_items_df = all_loadable_items_df.sort_values(by=['weight', 'quantity'], ascending=[False, False]).reset_index(drop=True)

    # Only flag items that have missing weight as a real issue
    missing_weight_df = merged_df[pd.isna(merged_df['weight'])].copy()
    if not missing_weight_df.empty:
        missing_weight_df['reason'] = 'Missing weight'

    frames = [
        missing_weight_df,
        unmatched_bom
    ]
    # Filter out empty or all-NaN DataFrames before concatenation
    frames = [f for f in frames if not f.empty and not f.isna().all(axis=None)]
    if frames:
        non_wall_or_missing_weight_df = pd.concat(frames, ignore_index=True)
    else:
        non_wall_or_missing_weight_df = pd.DataFrame()

    # --- Loading Simulation Functions ---
    def try_fit_trailers_combination(trailer_names, trailers_specs, materials_df, buffer):
        sorted_trailers = []
        # For Super B, expand to two separate trailers (sections)
        for t in trailer_names:
            if t == "Super B":
                # Assume Super B has two sections, divide specs accordingly
                specs = get_effective_trailer_dimensions(trailers_specs[t], buffer)
                # Divide weight and volume by 2 for each section
                for i in range(2):
                    sorted_trailers.append(f"Super B {i+1}")
            else:
                sorted_trailers.append(t)
        trailers_state = {}
        for tname in sorted_trailers:
            if tname.startswith("Super B"):
                # Get original Super B specs, divide weight/volume by 2
                specs_orig = get_effective_trailer_dimensions(trailers_specs["Super B"], buffer)
                specs = specs_orig.copy()
                specs['max_weight'] = specs_orig['max_weight'] / 2
                specs['max_volume'] = specs_orig['max_volume'] / 2
                # Optionally, you could also split length if needed, but we keep full dims for each section
            else:
                specs = get_effective_trailer_dimensions(trailers_specs[tname], buffer)
            trailers_state[tname] = {
                'specs': specs,
                'log': [],
                'total_volume': 0,
                'total_weight': 0,
                'total_height': 0,
                'loaded_set': set(),
                'stack_count': 0,
                'first_sheet': True,
                'remaining_width': specs['width'],
                'remaining_length': specs['length'],
                'stack_heights': [],
            }
        remaining_quantities = {idx: int(row['quantity']) for idx, row in materials_df.iterrows()}
        for tname in sorted_trailers:
            state = trailers_state[tname]
            specs = state['specs']
            sheets_per_row = specs['width'] // sheet_width
            max_rows = specs['length'] // sheet_length
            max_stack_height = specs['height']
            while state['remaining_length'] >= sheet_length:
                state['remaining_width'] = specs['width']
                while state['remaining_width'] >= sheet_width:
                    stack_max_height = max_stack_height * stack_height_limit_pct
                    temp_total_height = 0
                    temp_total_volume = state['total_volume']
                    temp_total_weight = state['total_weight']
                    temp_stack_count = 0
                    temp_first_sheet = True
                    loaded_any = False
                    current_stack_group = None  # Track current stack group
                    for assembly, group_df in materials_df.groupby('assembly'):
                        group_df_sorted = group_df.sort_values(by=['quantity', 'weight'], ascending=[False, False])
                        for idx, row in group_df_sorted.iterrows():
                            # --- Super B item length restriction ---
                            # Skip items longer than 350 inches for Super B sections
                            if tname.startswith("Super B") and row['length'] > 350:
                                continue
                            quantity_left = remaining_quantities.get(idx, 0)
                            if quantity_left <= 0:
                                continue
                            weight_per_sheet = row['weight']
                            loaded_count = 0
                            # Stacking logic with group separation
                            for i in range(quantity_left):
                                # If starting new stack, set current_stack_group
                                if current_stack_group is None:
                                    current_stack_group = row['group']
                                # If group changes, break to start new stack
                                if row['group'] != current_stack_group:
                                    break
                                if row['group'] == 'wall':
                                    added_height = 9.59 if temp_first_sheet else sheet_height
                                    added_volume = 29.42 if temp_first_sheet else sheet_volume
                                else:
                                    added_height = row['height']
                                    added_volume = row['volume']
                                temp_first_sheet = False
                                if (
                                    temp_total_height + added_height > stack_max_height or
                                    temp_total_volume + added_volume > specs['max_volume'] or
                                    temp_total_weight + weight_per_sheet > specs['max_weight']
                                ):
                                    break
                                temp_stack_count += 1
                                temp_total_height += added_height
                                temp_total_volume += added_volume
                                temp_total_weight += weight_per_sheet
                                loaded_count += 1
                            if loaded_count > 0:
                                state['total_volume'] = temp_total_volume
                                state['total_weight'] = temp_total_weight
                                state['log'].append((
                                    row['assembly'], row['material'], row['group'], "Loaded",
                                    loaded_count, weight_per_sheet * loaded_count, loaded_count * (sheet_volume if row['group'] == 'wall' else row['volume']),
                                    sheet_length if row['group'] == 'wall' else row['length'],
                                    sheet_width if row['group'] == 'wall' else row['width']
                                ))
                                state['loaded_set'].add((row['assembly'], row['material']))
                                remaining_quantities[idx] -= loaded_count
                                loaded_any = True
                        # After breaking due to group change, inner for-loop will move to next part
                    if temp_stack_count > 0:
                        state['stack_heights'].append(temp_total_height)
                    if not loaded_any:
                        break
                    state['remaining_width'] -= sheet_width + gap_between_stacks
                state['remaining_length'] -= sheet_length + gap_between_stacks_length
        unloaded_items = []
        # For enhanced unloaded logic
        # Build loaded assemblies by group and assembly for note
        loaded_assemblies_by_group_assembly = defaultdict(set)
        for tname, state in trailers_state.items():
            for entry in state['log']:
                if entry[3] == "Loaded":
                    loaded_assemblies_by_group_assembly[(entry[0], entry[2])].add(tname)
        for idx, row in materials_df.iterrows():
            quantity_left = remaining_quantities.get(idx, 0)
            if quantity_left > 0:
                weight_per_sheet = row['weight']
                item_length = sheet_length if row['group'] == 'wall' else row['length']
                item_width = sheet_width if row['group'] == 'wall' else row['width']
                item_weight = weight_per_sheet
                reasons = []
                # For each trailer, check which limits are exceeded
                for tname, state in trailers_state.items():
                    specs = state['specs']
                    avail_length = specs['length']
                    avail_width = specs['width']
                    avail_weight = specs['max_weight']
                    if item_length is not None and item_length > avail_length:
                        reasons.append(f"Too long for remaining space in trailer {tname}")
                    if item_width is not None and item_width > avail_width:
                        reasons.append(f"Too wide for remaining space in trailer {tname}")
                    if item_weight is not None and item_weight > avail_weight:
                        reasons.append(f"Too heavy for remaining capacity in trailer {tname}")
                reason = "; ".join(reasons) if reasons else "No suitable space remaining"
                # Note about partial loading
                note = None
                loaded_trailers = loaded_assemblies_by_group_assembly.get((row['assembly'], row['group']), set())
                if loaded_trailers:
                    note = f"Partially loaded in trailer(s): {', '.join(sorted(loaded_trailers))}"
                unloaded_items.append((
                    row['assembly'], row['material'], row['group'], reason,
                    quantity_left, weight_per_sheet * quantity_left, quantity_left * (sheet_volume if row['group'] == 'wall' else row['volume']),
                    item_length, item_width, note
                ))
        for idx, row in non_wall_or_missing_weight_df.iterrows():
            unloaded_items.append((
                row['assembly'], row['material'], row['group'],
                row['reason'],
                row.get('quantity', None), row.get('weight', None), None,
                None, None, None
            ))
        all_loaded = (len(unloaded_items) == 0 or all(
            (r[3].startswith("Material not loaded") for r in unloaded_items)
        ))
        return trailers_state, unloaded_items, all_loaded

    # --- Trailer Combination Search ---
    if country.lower() in restricted:
        all_trailer_names = [t for t in trailers if t != "Super B"]
    else:
        all_trailer_names = list(trailers.keys())
    all_combinations = []
    for r in range(1, 4):
        combis = list(itertools.combinations(all_trailer_names, r))
        all_combinations.extend(combis)
    best_config = None
    best_config_cost = None
    best_config_trailer_count = None
    best_config_state = None
    best_config_unloaded = None
    best_config_loaded_count = -1
    for combi in all_combinations:
        # Expand Super B into two trailers for internal logic
        expanded_combi = []
        for t in combi:
            if t == "Super B":
                expanded_combi.extend(["Super B 1", "Super B 2"])
            else:
                expanded_combi.append(t)
        state, unloaded, all_loaded = try_fit_trailers_combination(combi, trailers, all_loadable_items_df, buffer)
        # For cost, only count original trailers (not split Super B sections)
        total_cost = sum(cost_map[tname] for tname in combi)
        trailer_count = len(combi)
        loaded_count = sum(
            1 for tname in expanded_combi for entry in state[tname]['log'] if entry[3] == "Loaded"
        )
        if all_loaded:
            if best_config is None or total_cost < best_config_cost:
                best_config = combi
                best_config_cost = total_cost
                best_config_trailer_count = trailer_count
                best_config_state = state
                best_config_unloaded = unloaded
                best_config_loaded_count = loaded_count
        else:
            if best_config is None or (not best_config_state) or (
                loaded_count > best_config_loaded_count or
                (loaded_count == best_config_loaded_count and (total_cost < (best_config_cost or float('inf'))))
            ):
                best_config = combi
                best_config_cost = total_cost
                best_config_trailer_count = trailer_count
                best_config_state = state
                best_config_unloaded = unloaded
                best_config_loaded_count = loaded_count

    # --- Output Data Preparation ---
    log_with_trailer = []
    log_aggregator = defaultdict(lambda: [0, 0.0, 0.0])
    # For logging, expand Super B into its two sections
    expanded_best_config = []
    for t in best_config:
        if t == "Super B":
            expanded_best_config.extend(["Super B 1", "Super B 2"])
        else:
            expanded_best_config.append(t)
    for tname in expanded_best_config:
        for entry in best_config_state[tname]['log']:
            if entry[3] == "Loaded":
                key = (entry[0], entry[1], entry[2], tname)
                log_aggregator[key][0] += entry[4]
                log_aggregator[key][1] += entry[5]
                log_aggregator[key][2] += entry[6]
            else:
                log_with_trailer.append(entry + (None,))
    for (assembly, material, group, trailer), (qty, wt, vol) in log_aggregator.items():
        log_with_trailer.append((assembly, material, group, "Loaded", qty, wt, vol, trailer))
    assembly_material_trailers = defaultdict(set)
    for row in log_with_trailer:
        if row[3] == "Loaded":
            assembly_material_trailers[(row[0], row[1])].add(row[7])
    output_data = pd.DataFrame(log_with_trailer, columns=['Assembly', 'Material', 'Group', 'Status', 'Quantity', 'Weight', 'Volume', 'Trailer'])
    def split_trailer_note(row):
        if row['Status'] != "Loaded":
            return None
        trailers_ = assembly_material_trailers[(row['Assembly'], row['Material'])]
        if len(trailers_) > 1:
            other_trailers = sorted(t for t in trailers_ if t != row['Trailer'])
            return ", ".join(other_trailers)
        return None
    output_data['Split Assembly Trailers'] = output_data.apply(split_trailer_note, axis=1)
    used_trailers_summary = []
    # For summary, expand Super B into two separate sections
    for tname in expanded_best_config:
        if tname.startswith("Super B"):
            specs = get_effective_trailer_dimensions(trailers["Super B"], buffer)
            specs = specs.copy()
            specs['max_weight'] = specs['max_weight'] / 2
            specs['max_volume'] = specs['max_volume'] / 2
            cost = cost_map["Super B"] / 2  # For display, half the cost per section
        else:
            specs = get_effective_trailer_dimensions(trailers[tname], buffer)
            cost = cost_map[tname]
        state = best_config_state[tname]
        total_wt = state['total_weight']
        total_vol = state['total_volume']
        total_ht = state['total_height']
        weight_pct = round((total_wt / specs['max_weight']) * 100, 2) if total_wt is not None else None
        volume_pct = round((total_vol / specs['max_volume']) * 100, 2) if total_vol is not None else None
        height_pct = round((total_ht / specs['height']) * 100, 2) if total_ht is not None else None
        used_trailers_summary.append({
            'Trailer': tname,
            'Total Weight Used (lbs)': total_wt,
            'Total Volume Used (ft³)': total_vol,
            'Weight Used (%)': f"{weight_pct}%" if weight_pct is not None else None,
            'Volume Used (%)': f"{volume_pct}%" if volume_pct is not None else None,
            'Height Used (%)': f"{height_pct}%" if height_pct is not None else None,
            'Max Weight (lbs)': specs['max_weight'],
            'Max Volume (ft³)': specs['max_volume'],
            'Height Used (in)': total_ht,
            'Dimensions (LxWxH in)': f"{specs['length']}x{specs['width']}x{specs['height']}",
            'Total Cost ($)': cost,
            'Note': 'Selected Trailer'
        })
    unloaded_df = pd.DataFrame(best_config_unloaded, columns=['Assembly', 'Material', 'Group', 'Reason', 'Quantity', 'Weight', 'Volume', 'Length', 'Width', 'Note'])
    summary_df = pd.DataFrame(used_trailers_summary)

    # --- Post-processing: Remove columns and reset index as per instructions ---
    summary_df = summary_df.drop(columns=['Height Used (%)', 'Height Used (in)'], errors='ignore')
    summary_df.index = range(1, len(summary_df) + 1)
    output_data.index = range(1, len(output_data) + 1)
    unloaded_df.index = range(1, len(unloaded_df) + 1)
    try:
        floor_items_df.index = range(1, len(floor_items_df) + 1)
    except Exception:
        pass

    # --- Output to Temp Directory ---
    temp_dir = tempfile.mkdtemp(prefix="trailer_optimizer_")
    output_file = os.path.join(temp_dir, "Loading_Plan_Output.xlsx")
    with pd.ExcelWriter(output_file, engine='openpyxl') as writer:
        output_data.to_excel(writer, sheet_name='Loading Plan', index=False)
        summary_df.to_excel(writer, sheet_name='Summary', index=False)
        unloaded_df.to_excel(writer, sheet_name='Unloaded Items', index=False)
        stack_rows = []
        for tname in expanded_best_config:
            if tname.startswith("Super B"):
                specs = get_effective_trailer_dimensions(trailers["Super B"], buffer)
            else:
                specs = get_effective_trailer_dimensions(trailers[tname], buffer)
            stack_heights = best_config_state[tname].get('stack_heights', [])
            for i, height in enumerate(stack_heights, start=1):
                stack_rows.append({
                    'Trailer': tname,
                    'Stack Number': i,
                    'Stack Height (in)': round(height, 2),
                    'Stack Height Used (%)': f"{round((height/specs['height'])*100, 2)}%"
                })
        stack_df = pd.DataFrame(stack_rows)
        stack_df.to_excel(writer, sheet_name='Stack Breakdown', index=False)
    # --- Formatting ---
    import openpyxl
    wb = openpyxl.load_workbook(output_file)
    def auto_width(ws):
        for col in ws.columns:
            max_length = 0
            column = col[0].column_letter
            for cell in col:
                try:
                    if cell.value:
                        max_length = max(max_length, len(str(cell.value)))
                except:
                    pass
            adjusted_width = max_length + 2
            ws.column_dimensions[column].width = adjusted_width
    # Format for Loading Plan sheet
    ws1 = wb['Loading Plan']
    col_map = {cell.value: idx+1 for idx, cell in enumerate(ws1[1])}
    for row in ws1.iter_rows(min_row=2, max_row=ws1.max_row):
        if 'Weight' in col_map:
            cell = row[col_map['Weight']-1]
            if isinstance(cell.value, (int, float)):
                cell.number_format = '#,##0.00'
        if 'Volume' in col_map:
            cell = row[col_map['Volume']-1]
            if isinstance(cell.value, (int, float)):
                cell.number_format = '#,##0.00'
        if 'Quantity' in col_map:
            cell = row[col_map['Quantity']-1]
            cell.number_format = 'General'
        if 'Trailer' in col_map:
            cell = row[col_map['Trailer']-1]
            cell.number_format = '@'
    auto_width(ws1)
    ws2 = wb['Summary']
    col_map2 = {cell.value: idx+1 for idx, cell in enumerate(ws2[1])}
    for row in ws2.iter_rows(min_row=2, max_row=ws2.max_row):
        if 'Total Weight Used (lbs)' in col_map2:
            cell = row[col_map2['Total Weight Used (lbs)']-1]
            if isinstance(cell.value, (int, float)):
                cell.number_format = '#,##0.00'
        if 'Total Volume Used (ft³)' in col_map2:
            cell = row[col_map2['Total Volume Used (ft³)']-1]
            if isinstance(cell.value, (int, float)):
                cell.number_format = '#,##0.00'
        if 'Max Weight (lbs)' in col_map2:
            cell = row[col_map2['Max Weight (lbs)']-1]
            if isinstance(cell.value, (int, float)):
                cell.number_format = '#,##0'
        if 'Max Volume (ft³)' in col_map2:
            cell = row[col_map2['Max Volume (ft³)']-1]
            if isinstance(cell.value, (int, float)):
                cell.number_format = '#,##0'
        if 'Height Used (in)' in col_map2:
            cell = row[col_map2['Height Used (in)']-1]
            cell.number_format = 'General'
        if 'Dimensions (LxWxH in)' in col_map2:
            cell = row[col_map2['Dimensions (LxWxH in)']-1]
            cell.number_format = '@'
        if 'Total Cost ($)' in col_map2:
            cell = row[col_map2['Total Cost ($)']-1]
            if isinstance(cell.value, (int, float)):
                cell.number_format = '#,##0'
    auto_width(ws2)
    ws3 = wb['Unloaded Items']
    col_map3 = {cell.value: idx+1 for idx, cell in enumerate(ws3[1])}
    for row in ws3.iter_rows(min_row=2, max_row=ws3.max_row):
        if 'Weight' in col_map3:
            cell = row[col_map3['Weight']-1]
            if isinstance(cell.value, (int, float)):
                cell.number_format = '#,##0.00'
        if 'Volume' in col_map3:
            cell = row[col_map3['Volume']-1]
            if isinstance(cell.value, (int, float)):
                cell.number_format = '#,##0.00'
        if 'Quantity' in col_map3:
            cell = row[col_map3['Quantity']-1]
            cell.number_format = 'General'
        if 'Reason' in col_map3:
            cell = row[col_map3['Reason']-1]
            cell.number_format = '@'
    auto_width(ws3)
    ws4 = wb['Stack Breakdown']
    auto_width(ws4)
    wb.save(output_file)

    # --- Visualizations ---
    visuals = []
    # 1. Static PNGs
    def generate_trailer_layout_visuals(trailers_state, output_dir, buffer):
        os.makedirs(output_dir, exist_ok=True)
        paths = []
        for tname, state in trailers_state.items():
            # For Super B sections, use Super B specs for outer boundary
            if tname.startswith("Super B"):
                outer_specs = trailers["Super B"]
                specs = get_effective_trailer_dimensions(trailers["Super B"], buffer)
            else:
                outer_specs = trailers[tname]
                specs = get_effective_trailer_dimensions(trailers[tname], buffer)
            effective_specs = specs
            trailer_length = specs['length']
            trailer_width = specs['width']
            fig, ax = plt.subplots(figsize=(12, 6))
            ax.set_xlim(0, outer_specs['length'])
            ax.set_ylim(0, outer_specs['width'])
            ax.set_title(f"Trailer Layout: {tname}")
            ax.set_xlabel("Length (in)")
            ax.set_ylabel("Width (in)")
            ax.set_aspect('equal')
            ax.grid(True, linestyle="--", alpha=0.3)
            # Draw the outer original trailer boundary (full size)
            ax.add_patch(
                patches.Rectangle(
                    (0, 0), outer_specs['length'], outer_specs['width'],
                    linewidth=2, edgecolor='black', facecolor='none', linestyle='solid'
                )
            )
            # Draw the buffer zone rectangle inside the trailer (usable space after buffer)
            ax.add_patch(
                patches.Rectangle(
                    (buffer, buffer),
                    outer_specs['length'] - 2 * buffer,
                    outer_specs['width'] - 2 * buffer,
                    linewidth=1.5, edgecolor='red', facecolor='none', linestyle='dashed'
                )
            )
            x_offset = buffer
            y_offset = buffer
            stack_heights = state.get('stack_heights', [])
            loaded_logs = [log for log in state['log'] if log[3] == "Loaded"]
            stacks_to_display = []
            for i in range(min(len(stack_heights), len(loaded_logs))):
                stacks_to_display.append(loaded_logs[i])
            for stack_idx, height in enumerate(stack_heights):
                if stack_idx < len(stacks_to_display):
                    log = stacks_to_display[stack_idx]
                    assembly = log[0]
                    material = log[1]
                    qty = log[4]
                    # Use the actual length/width from the log for true stack size
                    length = log[7]
                    width = log[8]
                    group = log[2]
                else:
                    assembly = ""
                    material = ""
                    qty = ""
                    length = sheet_length
                    width = sheet_width
                    group = 'unknown'
                # Color assignment as per instructions
                if group == 'wall':
                    cmap = plt.get_cmap('Blues')
                    color = cmap(min(1.0, height / specs['height']))
                elif group == 'aerfloor':
                    color = (0.0, 0.6, 0.0, 0.9)
                elif group == 'roof':
                    cmap = plt.get_cmap('Oranges')
                    color = cmap(min(1.0, height / specs['height']))
                else:
                    cmap = plt.get_cmap('Greys')
                    color = cmap(min(1.0, height / specs['height']))
                rect = patches.Rectangle(
                    (x_offset, y_offset),
                    length,
                    width,
                    linewidth=1,
                    edgecolor='black',
                    facecolor=color,
                    alpha=0.9
                )
                ax.add_patch(rect)
                ax.text(
                    x_offset + length / 2,
                    y_offset + width / 2,
                    f"{'Stack ' + str(stack_idx+1)}\n{height:.2f} in",
                    ha='center', va='center', fontsize=6
                )
                y_offset += width + gap_between_stacks
                if y_offset + width > effective_specs['width'] + buffer:
                    y_offset = buffer
                    x_offset += length + gap_between_stacks_length
                if x_offset + length > effective_specs['length'] + buffer:
                    break
            plt.tight_layout()
            out_path = os.path.join(output_dir, f"{tname.replace(' ', '_')}_layout.png")
            plt.savefig(out_path)
            plt.close()
            paths.append(out_path)
        return paths
    # 2. Interactive 2D HTML
    def generate_interactive_2d_layouts(trailers_state, output_dir):
        import plotly.graph_objects as go
        os.makedirs(output_dir, exist_ok=True)
        paths = []
        for tname, state in trailers_state.items():
            if tname.startswith("Super B"):
                specs = get_effective_trailer_dimensions(trailers["Super B"], buffer)
            else:
                specs = get_effective_trailer_dimensions(trailers[tname], buffer)
            trailer_length = specs['length']
            trailer_width = specs['width']
            fig = go.Figure()
            x_offset = 0
            y_offset = 0
            stack_data = [
                {
                    "assembly": log[0],
                    "material": log[1],
                    "group": log[2],
                    "length": log[7],
                    "width": log[8],
                    "height": height
                }
                for log, height in zip(state["log"], state.get("stack_heights", []))
                if log[3] == "Loaded"
            ]
            for stack in stack_data:
                assembly = stack["assembly"]
                material = stack["material"]
                group = stack["group"]
                length = stack["length"]
                width = stack["width"]
                height = stack["height"]
                hover_text = f"Assembly: {assembly}<br>Material: {material}<br>Group: {group}<br>Height: {height:.2f} in"
                if group == "wall":
                    color = "blue"
                elif group == "roof":
                    color = "orange"
                elif group == "aerfloor":
                    color = "gold"
                else:
                    color = "gray"
                fig.add_shape(
                    type="rect",
                    x0=x_offset, y0=y_offset,
                    x1=x_offset+length, y1=y_offset+width,
                    line=dict(color="black"),
                    fillcolor=color,
                    opacity=0.7,
                )
                fig.add_trace(go.Scatter(
                    x=[x_offset + length/2],
                    y=[y_offset + width/2],
                    text=[hover_text],
                    mode="text",
                    showlegend=False,
                    hoverinfo="text"
                ))
                y_offset += width + gap_between_stacks
                if y_offset + width > trailer_width:
                    y_offset = 0
                    x_offset += length + gap_between_stacks_length
                if x_offset + length > trailer_length:
                    break
            fig.update_layout(
                title=f"Interactive 2D Trailer Layout: {tname}",
                xaxis=dict(range=[0, trailer_length], title="Length (in)", constrain='domain', scaleanchor='y', scaleratio=1),
                yaxis=dict(range=[0, trailer_width], title="Width (in)", constrain='domain'),
                width=900, height=500, margin=dict(l=0, r=0, t=40, b=0),
                plot_bgcolor="white",
            )
            out_path = os.path.join(output_dir, f"{tname.replace(' ', '_')}_interactive_2d.html")
            fig.write_html(out_path)
            paths.append(out_path)
        return paths
    # 3. Interactive 3D HTML
    def generate_interactive_3d_layouts(trailers_state, output_dir):
        import plotly.graph_objects as go
        os.makedirs(output_dir, exist_ok=True)
        paths = []
        def create_box(x, y, z, dx, dy, dz, group, hovertext):
            vertices = [
                [x, y, z],
                [x+dx, y, z],
                [x+dx, y+dy, z],
                [x, y+dy, z],
                [x, y, z+dz],
                [x+dx, y, z+dz],
                [x+dx, y+dy, z+dz],
                [x, y+dy, z+dz],
            ]
            faces = [
                [0,1,2], [0,2,3],
                [4,5,6], [4,6,7],
                [0,1,5], [0,5,4],
                [1,2,6], [1,6,5],
                [2,3,7], [2,7,6],
                [3,0,4], [3,4,7],
            ]
            x_vals, y_vals, z_vals = zip(*vertices)
            i, j, k = zip(*faces)
            if group == "wall":
                intensity = 0.2
                colorscale = [[0, 'rgb(173, 216, 230)'], [1, 'rgb(0, 0, 255)']]
            elif group == "roof":
                intensity = 0.8
                colorscale = [[0, 'rgb(255, 200, 0)'], [1, 'rgb(255, 100, 0)']]
            elif group == "aerfloor":
                intensity = 0.7
                colorscale = [[0, 'rgb(0, 200, 200)'], [1, 'rgb(0, 200, 0)']]
            else:
                intensity = 0.5
                colorscale = [[0, 'rgb(200, 200, 200)'], [1, 'rgb(100, 100, 100)']]
            return go.Mesh3d(
                x=x_vals, y=y_vals, z=z_vals,
                i=i, j=j, k=k,
                intensity=[intensity]*len(x_vals),
                colorscale=colorscale,
                cmin=0,
                cmax=1,
                opacity=0.75,
                hovertext=hovertext,
                hoverinfo="text",
                showscale=False,
                lighting=dict(ambient=0.5, diffuse=0.8, roughness=0.9, specular=0.2),
                flatshading=True
            )
        for tname, state in trailers_state.items():
            # For buffer zone, use outer_specs as the full trailer (not reduced by buffer)
            if tname.startswith("Super B"):
                outer_specs = trailers["Super B"]
                specs = get_effective_trailer_dimensions(trailers["Super B"], buffer)
            else:
                outer_specs = trailers[tname]
                specs = get_effective_trailer_dimensions(trailers[tname], buffer)
            trailer_length = specs['length']
            trailer_width = specs['width']
            trailer_height = specs['height']
            fig = go.Figure()
            # --- Add transparent buffer zone box before stacks ---
            buffer_length = outer_specs['length'] - 2 * buffer
            buffer_width = outer_specs['width'] - 2 * buffer
            buffer_height = outer_specs['height']
            # Only add buffer box if buffer > 0 and fits inside trailer
            if buffer_length > 0 and buffer_width > 0:
                # Vertices for the buffer box
                bx = [buffer, buffer+buffer_length, buffer+buffer_length, buffer, buffer, buffer+buffer_length, buffer+buffer_length, buffer]
                by = [buffer, buffer, buffer+buffer_width, buffer+buffer_width, buffer, buffer, buffer+buffer_width, buffer+buffer_width]
                bz = [0, 0, 0, 0, buffer_height, buffer_height, buffer_height, buffer_height]
                buffer_faces = [
                    [0,1,2], [0,2,3],    # bottom
                    [4,5,6], [4,6,7],    # top
                    [0,1,5], [0,5,4],    # front
                    [1,2,6], [1,6,5],    # right
                    [2,3,7], [2,7,6],    # back
                    [3,0,4], [3,4,7],    # left
                ]
                bi, bj, bk = zip(*buffer_faces)
                fig.add_trace(go.Mesh3d(
                    x=bx, y=by, z=bz,
                    i=bi, j=bj, k=bk,
                    color='red',
                    opacity=0.1,
                    hovertext="Buffer Zone",
                    hoverinfo="text",
                    showscale=False,
                    flatshading=True
                ))
            x_offset = 0
            y_offset = 0
            stack_data = [
                {
                    "assembly": log[0],
                    "material": log[1],
                    "group": log[2],
                    "length": log[7],
                    "width": log[8],
                    "height": height
                }
                for log, height in zip(state["log"], state.get("stack_heights", []))
                if log[3] == "Loaded"
            ]
            for stack in stack_data:
                assembly = stack["assembly"]
                material = stack["material"]
                group = stack["group"]
                length = stack["length"]
                width = stack["width"]
                height = stack["height"]
                hover_text = f"Assembly: {assembly}<br>Material: {material}<br>Group: {group}<br>Height: {height:.2f} in"
                mesh = create_box(
                    x_offset, y_offset, 0,
                    length, width, height,
                    group,
                    hover_text
                )
                fig.add_trace(mesh)
                y_offset += width + gap_between_stacks
                if y_offset + width > trailer_width:
                    y_offset = 0
                    x_offset += length + gap_between_stacks_length
                if x_offset + length > trailer_length:
                    break
            # Use full outer_specs for axis ranges and allow non-cubic aspect
            fig.update_layout(
                title=f"Trailer: {tname} - Interactive 3D Layout",
                scene=dict(
                    xaxis=dict(title='Length (in)', range=[0, outer_specs['length']]),
                    yaxis=dict(title='Width (in)', range=[0, outer_specs['width']]),
                    zaxis=dict(title='Height (in)', range=[0, outer_specs['height']]),
                    aspectmode='data'  # Keep real-world proportions
                ),
                margin=dict(l=0, r=0, b=0, t=40)
            )
            out_path = os.path.join(output_dir, f"{tname.replace(' ', '_')}_interactive_3d.html")
            fig.write_html(out_path)
            paths.append(out_path)
        return paths
    # --- Call and collect visualizations ---
    png_dir = os.path.join(temp_dir, "visuals")
    #html2d_dir = os.path.join(temp_dir, "interactive_visuals_2d")
    html3d_dir = os.path.join(temp_dir, "interactive_visuals_3d")
    visuals += generate_trailer_layout_visuals(best_config_state, png_dir, buffer)
    #visuals += generate_interactive_2d_layouts(best_config_state, html2d_dir)
    visuals += generate_interactive_3d_layouts(best_config_state, html3d_dir)
    # Add the Excel output file to visuals for completeness
    visuals.append(output_file)
    return summary_df, output_data, unloaded_df, visuals

